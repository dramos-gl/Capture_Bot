import os
import re
import shutil
import uuid
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from pypdf import PdfReader
from app.paths import BASE_DIR, TEMP_DIR

logger = logging.getLogger("OptimaCaptureBot.PhaseBExtractor")

# Expresión regular oficial del SAT para validar RFC
RFC_REGEX = re.compile(r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$", re.IGNORECASE)

def limpiar_razon_social(rs_raw: str) -> str:
    """
    Limpia la razón social removiendo comas, espacios excesivos y sufijos societarios comunes,
    dejando únicamente el nombre base del contribuyente.
    Ej: "CADURMA, S.A. DE C.V ." -> "CADURMA"
        "INMOCCIDENTE S.A. DE C.V." -> "INMOCCIDENTE"
    """
    if not rs_raw:
        return ""
    # Normalizar espacios múltiples a uno solo
    val = re.sub(r"\s+", " ", rs_raw).strip().upper()
    # Remover comas
    val = val.replace(",", "")
    # Remover sufijos societarios comunes con tolerancia de espacios y puntos
    pattern_societario = r"\bS\.?\s*A\.?\s*(?:B\.?)?\s+DE\s+C\.?\s*V\.?\b|\bS\.?\s*C\.?\b|\bS\.?\s*A\.?\b"
    val = re.sub(pattern_societario, "", val, flags=re.IGNORECASE)
    # Remover puntos residuales
    val = val.replace(".", "").strip()
    # Doble limpieza de espacios
    val = re.sub(r"\s+", " ", val).strip()
    return val

def extraer_datos_pdf(pdf_path: Path) -> dict:
    """
    Lee un archivo PDF y extrae los campos requeridos para el proceso:
    - Fecha alta
    - Referencia (17 dígitos)
    - RFC
    - Razón Social (limpia y sucia)
    - Código Postal (CP)
    - Importe
    - Fecha límite
    """
    try:
        reader = PdfReader(str(pdf_path))
        text = ""
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
        
        if not text.strip():
            raise ValueError("El archivo PDF está vacío o no contiene texto extraíble (podría ser una imagen escaneada).")

        # 1. Fecha alta
        fecha_alta_match = re.search(r"Fecha\s+alta:?\s*\n?\s*([\d\-\s\:]+)", text, re.IGNORECASE)
        fecha_alta_str = fecha_alta_match.group(1).strip() if fecha_alta_match else None
        
        # 2. Referencia (17 dígitos)
        ref_match = re.search(r"REFERENCIA:\s*(\d{17})", text, re.IGNORECASE)
        if not ref_match:
            # Búsqueda fallback por cualquier número de 17 dígitos
            ref_match = re.search(r"\b(\d{17})\b", text)
        referencia = ref_match.group(1).strip() if ref_match else None
        
        # 3. RFC
        rfc_match = re.search(r"Registro\s+Federal\s+de\s+Contribuyentes\s*\(RFC\):?\s*\n?\s*([A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3})", text, re.IGNORECASE)
        if not rfc_match:
            # Búsqueda fallback
            rfc_match = re.search(r"RFC:?\s*\n?\s*([A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3})", text, re.IGNORECASE)
        rfc = rfc_match.group(1).strip().upper() if rfc_match else None
        
        # 4. Razón Social
        rs_match = re.search(r"Apellido\s+Paterno,\s*Materno,\s*Nombres\(s\):?\s*\n?\s*([^\n]+)", text, re.IGNORECASE)
        rs_raw = rs_match.group(1).strip() if rs_match else ""
        rs_clean = limpiar_razon_social(rs_raw)
        
        # 5. Código Postal
        cp_match = re.search(r"Codigo\s+Postal:?\s*\n?\s*(\d{5})", text, re.IGNORECASE)
        cp = cp_match.group(1).strip() if cp_match else None
        
        # 6. Importe
        importe_match = re.search(r"IMPORTE:\s*\$?\s*([\d,]+(?:\.\d{2})?)", text, re.IGNORECASE)
        if not importe_match:
            # Fallback en total a pagar
            importe_match = re.search(r"TOTAL\s+A\s+PAGAR:?\s*\$?\s*([\d,]+(?:\.\d{2})?)", text, re.IGNORECASE)
        importe_str = importe_match.group(1).strip() if importe_match else None
        
        # Intentar convertir importe a float
        importe = None
        if importe_str:
            try:
                importe = float(importe_str.replace(",", ""))
            except ValueError:
                importe = importe_str

        # 7. Fecha Límite (estricto YYYY-MM-DD para evitar basuras del pie de página)
        fecha_limite_match = re.search(r"FECHA\s+LIMITE:\s*(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
        fecha_limite = fecha_limite_match.group(1).strip() if fecha_limite_match else None
        
        return {
            "fecha_alta": fecha_alta_str,
            "referencia": referencia,
            "rfc": rfc,
            "razon_social_raw": rs_raw,
            "razon_social": rs_clean,
            "cp": cp,
            "importe": importe,
            "fecha_limite": fecha_limite,
            "success": bool(referencia and rfc and rs_clean and cp)
        }
    except Exception as e:
        logger.error(f"Error procesando {pdf_path.name}: {e}")
        return {
            "success": False,
            "error": str(e)
        }

def procesar_carpeta_fase_b(
    origen_dir: str, 
    destino_dir: str, 
    conservar_originales: bool = True,
    log_callback=None,
    progress_callback=None
) -> dict:
    """
    Función principal de la Fase B.
    1. Lee todos los PDFs en origen_dir.
    2. Extrae sus datos y maneja errores moviendo corruptos/incompletos.
    3. Agrupa por RFC y ordena cronológicamente por 'Fecha alta'.
    4. Copia y renombra PDFs a 'Destino/[RFC]/Referencia_RFC_id.pdf'.
    5. Genera libros de Excel particionados a 300 registros.
    """
    def log(msg, nivel="INFO"):
        logger.log(logging.INFO if nivel == "INFO" else logging.WARNING, msg)
        if log_callback:
            log_callback(f"[FASE B] {msg}", nivel)

    path_origen = Path(origen_dir).resolve()
    path_destino = Path(destino_dir).resolve()
    
    if not path_origen.is_dir():
        log(f"El directorio de origen no existe: {origen_dir}", "ERROR")
        return {"success": False, "message": "Directorio de origen no válido."}

    # Asegurar la creación de las carpetas de destino base
    path_destino.mkdir(parents=True, exist_ok=True)
    
    # 1. Buscar todos los archivos PDF
    pdf_files = list(path_origen.rglob("*.pdf"))
    total_archivos = len(pdf_files)
    
    if total_archivos == 0:
        log("No se encontraron archivos PDF en la carpeta seleccionada.", "WARNING")
        return {"success": True, "processed": 0, "message": "No se encontraron PDFs."}
        
    log(f"Iniciando escaneo de {total_archivos} archivos PDF...")
    
    registros_validos = []
    
    # Crear carpetas de contingencia
    dir_errores = path_destino / "Errores_Lectura"
    dir_incompletos = path_destino / "Incompletos_o_Invalidos"
    
    procesados = 0
    for pdf_path in pdf_files:
        procesados += 1
        if progress_callback:
            progress_callback(procesados, total_archivos, f"Procesando {pdf_path.name}")
            
        datos = extraer_datos_pdf(pdf_path)
        datos["original_path"] = pdf_path
        
        if not datos["success"]:
            # Manejo de error de lectura o campos críticos vacíos
            razon_falla = datos.get("error", "Faltan campos críticos obligatorios (Referencia/RFC/CP).")
            log(f"Inconsistencia en {pdf_path.name}: {razon_falla}", "WARNING")
            
            # Decidir carpeta destino
            target_dir = dir_errores if "error" in datos else dir_incompletos
            target_dir.mkdir(parents=True, exist_ok=True)
            
            # Copiar o mover
            shutil.copy2(pdf_path, target_dir / pdf_path.name)
            if not conservar_originales:
                try:
                    pdf_path.unlink()
                except Exception as ex:
                    log(f"No se pudo eliminar el original de {pdf_path.name}: {ex}", "WARNING")
            continue
            
        registros_validos.append(datos)
        
    total_validos = len(registros_validos)
    log(f"Escaneo finalizado. PDFs válidos: {total_validos} | PDFs con error/incompletos: {total_archivos - total_validos}")
    
    if total_validos == 0:
        log("No se obtuvieron registros válidos para generar archivos de Excel.", "WARNING")
        return {"success": True, "processed": 0, "message": "Cero registros válidos."}

    # 2. Agrupar por RFC
    grupos_rfc = {}
    for reg in registros_validos:
        rfc_val = reg["rfc"]
        if rfc_val not in grupos_rfc:
            grupos_rfc[rfc_val] = []
        grupos_rfc[rfc_val].append(reg)
        
    # Ruta de la plantilla base
    template_path = BASE_DIR / "Optima_Capture_Bot.xlsx"
    if not template_path.exists():
        log(f"No se encuentra la plantilla Excel en {template_path}. Se buscará en raíz.", "WARNING")
        template_path = Path("Optima_Capture_Bot.xlsx")
        
    if not template_path.exists():
        log("Error crítico: No se encontró la plantilla base 'Optima_Capture_Bot.xlsx'. Asegúrese de colocarla en el directorio raíz.", "ERROR")
        return {"success": False, "message": "Falta plantilla Excel."}

    # 3. Procesar cada grupo RFC
    for rfc_key, lote_registros in grupos_rfc.items():
        # Crear subcarpeta para el RFC
        dir_rfc = path_destino / rfc_key
        dir_rfc.mkdir(parents=True, exist_ok=True)
        
        # Ordenar por fecha de alta (ascendente)
        # Manejador de fechas robusto ante formatos inesperados
        def parse_date_sort(item):
            d_str = item.get("fecha_alta")
            if not d_str:
                return datetime.min
            try:
                # Intentar formato normal
                return datetime.strptime(d_str.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    return datetime.strptime(d_str.strip(), "%Y-%m-%d")
                except ValueError:
                    return datetime.min

        lote_registros.sort(key=parse_date_sort)
        
        # Asignar ID consecutivo
        for index, reg in enumerate(lote_registros, start=1):
            reg["id"] = index

        total_grupo = len(lote_registros)
        log(f"Procesando RFC '{rfc_key}': {total_grupo} registros ordenados cronológicamente.")
        
        # 4. Copiar y renombrar PDFs
        for reg in lote_registros:
            ref = reg["referencia"]
            idx = reg["id"]
            orig = reg["original_path"]
            
            # Nombre nuevo: Referencia_RFC_id.pdf
            nuevo_nombre = f"{ref}_{rfc_key}_{idx}.pdf"
            dest_pdf_path = dir_rfc / nuevo_nombre
            
            try:
                shutil.copy2(orig, dest_pdf_path)
                if not conservar_originales:
                    orig.unlink()
            except Exception as e:
                log(f"Error al copiar/renombrar PDF {orig.name} -> {nuevo_nombre}: {e}", "WARNING")

        # 5. Generar Excel particionados en bloques de 300 registros
        tamano_particion = 300
        total_particiones = (total_grupo + tamano_particion - 1) // tamano_particion
        
        for part in range(1, total_particiones + 1):
            inicio = (part - 1) * tamano_particion
            fin = min(part * tamano_particion, total_grupo)
            subset = lote_registros[inicio:fin]
            
            # Cargar libro plantilla
            wb = openpyxl.load_workbook(str(template_path))
            
            # 5.1 Escribir Catálogo de RFC (Único registro en fila 2)
            if "Catalogo_RFC" in wb.sheetnames:
                ws_cat = wb["Catalogo_RFC"]
                # Limpiar cualquier contenido de datos previo en fila 2
                for col in range(1, 10):
                    ws_cat.cell(row=2, column=col).value = None
                
                # Escribir contribuyente
                first_reg = subset[0]
                ws_cat.cell(row=2, column=1, value=rfc_key)
                ws_cat.cell(row=2, column=2, value=first_reg["razon_social"])
                
                # CP como número si es posible
                cp_val = first_reg["cp"]
                try:
                    ws_cat.cell(row=2, column=3, value=int(cp_val))
                except (ValueError, TypeError):
                    ws_cat.cell(row=2, column=3, value=cp_val)
            else:
                log("Advertencia: Plantilla no contiene pestaña 'Catalogo_RFC'.", "WARNING")

            # 5.2 Escribir Control de Referencias
            if "Control_Referencias" in wb.sheetnames:
                ws_ref = wb["Control_Referencias"]
                
                # Limpiar filas existentes debajo de cabeceras (fila 2 en adelante)
                # openpyxl max_row puede ser mayor si hay formatos vacíos
                for r in range(2, ws_ref.max_row + 1):
                    for c in range(1, 15):
                        ws_ref.cell(row=r, column=c).value = None
                
                # Escribir registros
                for idx, reg in enumerate(subset, start=2):
                    ws_ref.cell(row=idx, column=1, value=reg["id"])                 # Col A (ID)
                    ws_ref.cell(row=idx, column=2, value=reg["referencia"])         # Col B (Referencia)
                    ws_ref.cell(row=idx, column=3).value = None                     # Col C (RFC vacío)
                    ws_ref.cell(row=idx, column=4, value="PENDIENTE")               # Col D (Estado)
                    ws_ref.cell(row=idx, column=5).value = None                     # Col E (Lote vacío)
                    ws_ref.cell(row=idx, column=6).value = None                     # Col F (Fecha ejecución)
                    ws_ref.cell(row=idx, column=7).value = None                     # Col G (Error vacío)
                    ws_ref.cell(row=idx, column=8, value=1)                         # Col H (Cantidad = 1)
                    ws_ref.cell(row=idx, column=9, value=reg["importe"])            # Col I (Importe)
                    ws_ref.cell(row=idx, column=10).value = None                    # Col J (Porcentaje vacío)
                    
                    # Columna K: Agregar encabezado si no existe en fila 1, escribir Fecha Límite
                    ws_ref.cell(row=1, column=11, value="FECHA LIMITE")
                    ws_ref.cell(row=idx, column=11, value=reg["fecha_limite"])
            else:
                log("Error: La plantilla no contiene la pestaña obligatoria 'Control_Referencias'.", "ERROR")
                wb.close()
                continue
                
            # Guardado atómico del libro de Excel
            excel_filename = f"{rfc_key}_OCB_{part}.xlsx"
            excel_dest_path = dir_rfc / excel_filename
            
            try:
                # Verificar directorio temp
                if not TEMP_DIR.exists():
                    TEMP_DIR.mkdir(parents=True, exist_ok=True)
                
                tmp_excel_path = TEMP_DIR / f"tmp_{uuid.uuid4().hex}.xlsx"
                wb.save(str(tmp_excel_path))
                wb.close()
                
                if tmp_excel_path.exists():
                    shutil.move(str(tmp_excel_path), str(excel_dest_path))
                    log(f"Excel generado con éxito: {excel_dest_path.name} (Registros {inicio + 1} al {fin})")
            except Exception as save_err:
                log(f"Error al guardar Excel {excel_filename}: {save_err}", "ERROR")

    log("=== PROCESAMIENTO DE FASE B COMPLETADO CON ÉXITO ===")
    return {"success": True, "processed": total_validos, "message": "Proceso completado."}
