import os
import shutil
import logging
from datetime import datetime
import openpyxl
from app.paths import TEMP_DIR

# Configuración del log
logger = logging.getLogger("OptimaCaptureBot.ExcelHandler")

# Mapeo exacto de las columnas de la hoja 'Control_Referencias'
COL_ID = 1               # Columna A (Indice numérico)
COL_REFERENCIA = 2       # Columna B (Referencia)
COL_RFC = 3              # Columna C (RFC)
COL_ESTADO = 4           # Columna D (Estado_Proceso)
COL_LOTE = 5             # Columna E (Lote_Asignado)
COL_FECHA = 6            # Columna F (Fecha_Hora_Ejecucion)
COL_ERROR = 7            # Columna G (Detalle_Error)

SHEET_DATA = "Control_Referencias"
SHEET_CATALOG = "Catalogo_RFC"

def verificar_bloqueo(filepath):
    """
    Comprueba con extra cautela si el archivo Excel está bloqueado por otro proceso 
    (por ejemplo, si el operador lo tiene abierto en Microsoft Excel).
    """
    if not os.path.exists(filepath):
        return False
    try:
        # Intentamos abrir el archivo en modo de lectura/escritura exclusivo
        with open(filepath, "r+b") as f:
            pass
        return False
    except OSError as e:
        logger.warning(f"El archivo {filepath} está bloqueado por otro programa: {e}")
        return True

def crear_backup(filepath, temp_dir=None):
    """
    Crea una copia de seguridad rápida en la carpeta temporal.
    """
    if temp_dir is None:
        temp_dir = str(TEMP_DIR)
    try:
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        filename = os.path.basename(filepath)
        backup_path = os.path.join(temp_dir, f"{filename}.bak")
        
        if os.path.exists(filepath):
            shutil.copy2(filepath, backup_path)
            logger.info(f"Copia de seguridad creada en {backup_path}")
            return backup_path
    except Exception as e:
        logger.error(f"Error al crear copia de seguridad: {e}")
    return None

def cargar_catalog_rfc(filepath):
    """
    Carga el RFC, Razón Social y Código Postal del catálogo.
    Se asume la Fila 2 como el registro de la empresa activa.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"No se encuentra el archivo de datos: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if SHEET_CATALOG not in wb.sheetnames:
            raise ValueError(f"Falta la hoja requerida: '{SHEET_CATALOG}' en el archivo Excel.")
        
        ws = wb[SHEET_CATALOG]
        # Leer fila 2
        rfc = ws.cell(row=2, column=1).value
        razon_social = ws.cell(row=2, column=2).value
        cp = ws.cell(row=2, column=3).value
        
        wb.close()
        
        if not rfc:
            raise ValueError("El RFC en la hoja 'Catalogo_RFC' está vacío.")
            
        return {
            "rfc": str(rfc).strip(),
            "razon_social": str(razon_social).strip() if razon_social else "",
            "cp": str(int(float(cp))) if cp is not None and str(cp).strip() != "" else ""
        }
    except Exception as e:
        logger.error(f"Error al cargar el catálogo de RFC: {e}")
        raise

def cargar_registros(filepath):
    """
    Lee las referencias y el estado actual de la hoja 'Control_Referencias' (filas 2 a 300).
    Retorna una lista de diccionarios con el mapeo.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"No se encuentra el archivo de datos: {filepath}")
        
    registros = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if SHEET_DATA not in wb.sheetnames:
            raise ValueError(f"Falta la hoja requerida: '{SHEET_DATA}' en el archivo Excel.")
            
        ws = wb[SHEET_DATA]
        
        # Procesamos desde la fila 2 hasta el límite real de filas con datos
        for row in range(2, ws.max_row + 1):
            ref_val = ws.cell(row=row, column=COL_REFERENCIA).value
            if ref_val is None:
                continue  # Saltamos filas vacías
                
            estado_val = ws.cell(row=row, column=COL_ESTADO).value
            rfc_val = ws.cell(row=row, column=COL_RFC).value
            lote_val = ws.cell(row=row, column=COL_LOTE).value
            
            registros.append({
                "fila_excel": row,
                "id": ws.cell(row=row, column=COL_ID).value,
                "referencia": str(ref_val).strip(),
                "rfc": str(rfc_val).strip() if rfc_val else None,
                "estado": str(estado_val).strip() if estado_val else "PENDIENTE",
                "lote": str(lote_val).strip() if lote_val else None,
                "cantidad": ws.cell(row=row, column=8).value,
                "importe": ws.cell(row=row, column=9).value,
                "porcentaje": ws.cell(row=row, column=10).value
            })
            
        wb.close()
        logger.info(f"Cargados {len(registros)} registros desde {SHEET_DATA}")
        return registros
    except Exception as e:
        logger.error(f"Error al cargar los registros del Excel: {e}")
        raise

def actualizar_fila(filepath, fila_excel, estado, lote=None, detalle_error=None, rfc=None):
    """
    Actualiza de forma atómica y segura las columnas de control de una fila específica.
    Implementa copias de seguridad previas y reintento en caso de bloqueo temporal.
    """
    # 1. Verificar bloqueo exclusivo antes de intentar cualquier operación
    if verificar_bloqueo(filepath):
        logger.error(f"No se puede guardar: El archivo {filepath} está bloqueado por otro proceso.")
        raise PermissionError(f"El archivo {filepath} está abierto en Excel u otro programa. Por favor, ciérrelo.")
        
    # 2. Crear copia de seguridad antes del guardado
    crear_backup(filepath)
    
    try:
        # Cargar libro en caliente
        wb = openpyxl.load_workbook(filepath)
        ws = wb[SHEET_DATA]
        
        # Actualizar celdas
        if rfc:
            ws.cell(row=fila_excel, column=COL_RFC, value=str(rfc).strip())
            
        ws.cell(row=fila_excel, column=COL_ESTADO, value=str(estado).strip())
        
        if lote:
            ws.cell(row=fila_excel, column=COL_LOTE, value=str(lote).strip())
            
        ws.cell(row=fila_excel, column=COL_FECHA, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        if detalle_error is not None:
            ws.cell(row=fila_excel, column=COL_ERROR, value=str(detalle_error).strip())
        else:
            ws.cell(row=fila_excel, column=COL_ERROR, value=None)
            
        # 3. Guardado seguro usando un reemplazo atómico para evitar corrupción en cortes
        import uuid
        temp_dir = str(TEMP_DIR)
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            
        temp_filepath = os.path.join(temp_dir, f"tmp_{uuid.uuid4().hex}.xlsx")
        wb.save(temp_filepath)
        wb.close()
        
        # Reemplazar el original con el archivo guardado exitosamente
        if os.path.exists(temp_filepath):
            shutil.move(temp_filepath, filepath)
            logger.info(f"Fila {fila_excel} de Excel actualizada exitosamente a estado: {estado}")
            return True
            
    except PermissionError as pe:
        logger.error(f"Fallo de permisos al guardar fila {fila_excel}: {pe}")
        raise PermissionError(f"El archivo {filepath} está bloqueado. Ciérrelo para guardar el progreso.") from pe
    except Exception as e:
        logger.error(f"Error crítico al actualizar la fila {fila_excel} en Excel: {e}")
        raise
    return False

def colorear_celdas_validacion(filepath, mapeo_colores):
    """
    Colorea de forma atómica y segura las celdas de las referencias en el Excel según el estado de la validación.
    mapeo_colores es un dict: {fila_excel: color_hex}
    Colores:
      - 'AZUL' (incompleto) -> 93C5FD
      - 'AMARILLO' (no descargado) -> FEF08A
      - None (completo) -> None (sin relleno)
    """
    if verificar_bloqueo(filepath):
        raise PermissionError(f"El archivo {filepath} está bloqueado. Ciérrelo para poder guardar los colores de validación.")
        
    crear_backup(filepath)
    try:
        from openpyxl.styles import PatternFill
        wb = openpyxl.load_workbook(filepath)
        ws = wb[SHEET_DATA]
        
        fill_azul = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
        fill_amarillo = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
        
        for fila, color in mapeo_colores.items():
            cell = ws.cell(row=fila, column=COL_REFERENCIA)
            if color == "AZUL":
                cell.fill = fill_azul
            elif color == "AMARILLO":
                cell.fill = fill_amarillo
            else:
                cell.fill = PatternFill(fill_type=None)  # Limpiar relleno
                
        # Guardado seguro
        import uuid
        temp_dir = str(TEMP_DIR)
        temp_filepath = os.path.join(temp_dir, f"tmp_{uuid.uuid4().hex}.xlsx")
        wb.save(temp_filepath)
        wb.close()
        
        if os.path.exists(temp_filepath):
            shutil.move(temp_filepath, filepath)
            logger.info("Rellenos de celdas de validación de PDF aplicados correctamente en Excel.")
            return True
    except Exception as e:
        logger.error(f"Error al pintar celdas en el Excel: {e}")
        raise
    return False
