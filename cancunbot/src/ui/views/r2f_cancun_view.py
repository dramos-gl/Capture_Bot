"""
R2F-Cancún — Unified Dashboard View (Standalone Window)
Permite gestionar la descarga de Recibos y la posterior generación de Facturas
en una interfaz unificada mediante un interruptor de modo (CustomSwitch).
"""
import datetime
import os
import logging
from pathlib import Path
from typing import Optional

from PySide6.QtWidgets import (
    QWidget, QMainWindow, QVBoxLayout, QHBoxLayout, QGridLayout, 
    QFrame, QLabel, QPushButton, QTextEdit, QMessageBox, 
    QProgressBar, QMenu, QFileDialog, QTabWidget, QDialog
)
from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtGui import QFont, QCursor

# Reutilizar componentes del Design System del SAR
from sar.src.ui.design_system.components import (
    CustomCard, StyledDataTable, CustomButton, CustomLabel, 
    CustomCheckBox, CustomSwitch, MetricBox
)
from sar.src.ui.design_system.components.atoms.gl_status_indicator import GLStatusIndicator
from sar.src.ui.design_system.tokens.colors import Colors
from sar.src.storage.repositories import ConfigRepository

from cancunbot.src.storage.cancunbot_repos import LoteFolioRepository, FolioCancunRepository, ReciboCancunRepository
from cancunbot.src.core.bot_recibo_worker import BotReciboCunWorker

logger = logging.getLogger(__name__)


class R2FCancunWindow(QMainWindow):
    """MainWindow wrapper que cumple con el patrón de ventanas independientes del SAR (ej: BillingBotWindow)."""
    logout_requested = Signal()

    def __init__(self, db_connector, sesion_id, usuario_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("R2F-Cancún (Recibos y Facturación)")
        self.resize(1100, 750)
        
        self.view = R2FCancunView(db_connector, sesion_id, usuario_id, self)
        self.setCentralWidget(self.view)
        
        # Propagar señal de logout
        self.view.logout_requested.connect(self.logout_requested.emit)

    def closeEvent(self, event):
        self.view.closeEvent(event)


class R2FCancunView(QWidget):
    """Vista del dashboard unificado con selector de modo (Recibos / Facturas)."""
    logout_requested = Signal()

    def __init__(self, db_connector, sesion_id, usuario_id, parent=None):
        super().__init__(parent)
        self.db_connector = db_connector
        self.sesion_id = sesion_id
        self.usuario_id = usuario_id
        
        self.active_worker = None
        self.selected_lote_id = None
        
        # Layout principal
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(10, 10, 10, 10)
        self.main_layout.setSpacing(10)

        # Cargar ruta base configurada
        self.default_output_dir = "T:\\CANCUN"
        try:
            with self.db_connector.get_session() as session:
                repo = ConfigRepository(session)
                db_dir = repo.get_parametro("CANCUN_PDF_BASE_PATH")
                if db_dir:
                    self.default_output_dir = db_dir
        except Exception as e:
            logger.error(f"Error cargando directorio de Cancún: {e}")

        self.selected_custom_path = None

        # Construir paneles de UI
        self._build_header()
        self._build_controls_and_metrics()
        self._build_tables_panel()
        self._build_console_panel()

        # Cargar datos iniciales
        self._refresh_lotes_table()
        self._verify_paths()

    def _build_header(self):
        header_frame = QFrame()
        header_frame.setObjectName("botHeaderFaceA")  # Reusar estilos CSS del SAR
        h_layout = QHBoxLayout(header_frame)
        h_layout.setContentsMargins(0, 0, 0, 0)
        h_layout.setSpacing(12)

        self.lbl_titulo = QLabel("📥 R2F-CANCÚN — MODO: RECIBOS")
        # Cambiado el color a blanco (#ffffff) para contraste total con el fondo oscuro del Header del bot.
        self.lbl_titulo.setStyleSheet("font-size: 16px; font-weight: bold; color: #ffffff;")
        h_layout.addWidget(self.lbl_titulo)
        h_layout.addStretch()

        # Switch de Modo (RECIBOS / FACTURAS)
        self.switch_modo = CustomSwitch("MODO FACTURAS")
        self.switch_modo.setChecked(False)
        self.switch_modo.toggled.connect(self._on_modo_toggled)
        h_layout.addWidget(self.switch_modo)

        # Indicador de estado del portal
        self.lbl_portal_status = QLabel("Portal: INACTIVO")
        self.lbl_portal_status.setStyleSheet("background-color: #334155; padding: 4px 12px; border-radius: 12px; font-size: 12px; color: white;")
        h_layout.addWidget(self.lbl_portal_status)

        # Botones de Gear y User
        self.btn_gear = QPushButton("⚙")
        self.btn_gear.setObjectName("iconHeaderBtn")
        self.btn_gear.clicked.connect(self._on_gear_clicked)
        h_layout.addWidget(self.btn_gear)

        self.btn_user = QPushButton("👤")
        self.btn_user.setObjectName("iconHeaderBtn")
        self.btn_user.clicked.connect(self._on_user_clicked)
        h_layout.addWidget(self.btn_user)

        self.main_layout.addWidget(header_frame)

    def _build_controls_and_metrics(self):
        top_layout = QHBoxLayout()
        top_layout.setSpacing(10)

        # Panel de Controles
        controles_frame = QFrame()
        controles_frame.setObjectName("card")
        c_layout = QVBoxLayout(controles_frame)
        c_layout.setContentsMargins(8, 8, 8, 8)

        lbl_c = CustomLabel("⚙ CONTROLES", variant="subheader")
        c_layout.addWidget(lbl_c)

        self.chk_autonomo = CustomSwitch("🤖 Modo Autónomo (Visible)")
        self.chk_autonomo.setChecked(True)  # Activo por defecto (Visible en Playwright)
        self.chk_autonomo.setEnabled(False) # Estático, no editable por el usuario
        c_layout.addWidget(self.chk_autonomo)

        self.btn_importar_excel = CustomButton("📁 Importar Excel de Folios", is_secondary=True)
        self.btn_importar_excel.clicked.connect(self._on_importar_excel)
        c_layout.addWidget(self.btn_importar_excel)

        self.btn_descargar_plantilla = CustomButton("⬇ Descargar Plantilla Excel", is_secondary=True)
        self.btn_descargar_plantilla.clicked.connect(self._on_descargar_plantilla)
        c_layout.addWidget(self.btn_descargar_plantilla)

        # Replicando selector de ruta de descarga de Bot Face A
        lbl_path_title = CustomLabel("📁 Ruta de Descarga / Almacenamiento:", variant="body")
        lbl_path_title.setStyleSheet("font-weight: bold;")
        c_layout.addWidget(lbl_path_title)
        
        path_input_layout = QHBoxLayout()
        display_label_text = f"Por defecto ({self.default_output_dir})"
        self.lbl_download_path_display = CustomLabel(display_label_text, variant="body")
        self.lbl_download_path_display.setStyleSheet("background-color: #f9fafb; padding: 6px; border: 1px solid #d1d5db; border-radius: 4px; font-size: 11px;")
        path_input_layout.addWidget(self.lbl_download_path_display, stretch=4)
        
        self.btn_browse = CustomButton("...", is_secondary=True)
        self.btn_browse.setObjectName("secondaryBtn")
        self.btn_browse.setStyleSheet("padding: 4px 8px; font-weight: bold;")
        self.btn_browse.clicked.connect(self._on_browse_path_clicked)
        path_input_layout.addWidget(self.btn_browse, stretch=1)
        c_layout.addLayout(path_input_layout)

        # GLStatusIndicator pill under the download path
        self.status_indicator = GLStatusIndicator()
        c_layout.addWidget(self.status_indicator)

        self.btn_iniciar = CustomButton("▶ Iniciar Bot", is_secondary=False)
        self.btn_iniciar.setObjectName("primaryBtn")
        self.btn_iniciar.clicked.connect(self._on_iniciar_bot)
        c_layout.addWidget(self.btn_iniciar)

        self.btn_detener = CustomButton("🛑 Detener Bot", is_secondary=True)
        self.btn_detener.setObjectName("dangerBtn")
        self.btn_detener.setEnabled(False)
        self.btn_detener.clicked.connect(self._on_detener_bot)
        c_layout.addWidget(self.btn_detener)

        top_layout.addWidget(controles_frame, stretch=1)

        # Panel de Métricas
        metricas_frame = QFrame()
        metricas_frame.setObjectName("card")
        m_layout = QVBoxLayout(metricas_frame)
        m_layout.setContentsMargins(8, 8, 8, 8)

        lbl_m = CustomLabel("📊 ESTADO DEL PROCESO", variant="subheader")
        m_layout.addWidget(lbl_m)

        grid_m = QGridLayout()
        self.box_pendientes = MetricBox("Pendientes", "0", "#3b82f6")
        self.box_exitosos = MetricBox("Exitosos", "0", "#10b981")
        self.box_errores = MetricBox("Errores", "0", "#ef4444")

        grid_m.addWidget(self.box_pendientes, 0, 0)
        grid_m.addWidget(self.box_exitosos, 0, 1)
        grid_m.addWidget(self.box_errores, 0, 2)

        self.lbl_lote_actual_info = CustomLabel("Lote seleccionado: Ninguno", variant="muted")
        grid_m.addWidget(self.lbl_lote_actual_info, 1, 0, 1, 3)

        m_layout.addLayout(grid_m)
        top_layout.addWidget(metricas_frame, stretch=2)

        self.main_layout.addLayout(top_layout)

    def _build_tables_panel(self):
        self.tables_card = CustomCard("📑 BANDEJA DE TRABAJO")
        # CustomCard ya tiene un container interno con layout vertical.
        # Agregamos el tab_widget usando .add_widget() en lugar de asignarle un layout directamente a la tarjeta.
        self.tab_widget = QTabWidget()
        
        # Tabla de Lotes
        self.table_lotes = StyledDataTable([
            "ID", "Folio Lote", "Origen", "Total Folios", "Procesados", "Estado"
        ])
        self.table_lotes.doubleClicked.connect(self._on_lote_double_clicked)
        self.tab_widget.addTab(self.table_lotes, "Lotes de Folios")

        # Tabla de Detalles (Folios / Recibos)
        self.table_detalles = StyledDataTable([
            "ID", "Folio/Referencia", "Tipo", "Intentos", "Estado"
        ])
        self.tab_widget.addTab(self.table_detalles, "Folios en Lote")

        self.tables_card.add_widget(self.tab_widget)
        self.main_layout.addWidget(self.tables_card, stretch=2)

    def _build_console_panel(self):
        console_card = CustomCard("📡 BITÁCORA DE OPERACIÓN")

        self.txt_console = QTextEdit()
        self.txt_console.setReadOnly(True)
        self.txt_console.setObjectName("console")
        self.txt_console.setStyleSheet("font-family: 'Consolas', 'Courier New', monospace; font-size: 11px; background-color: #1e293b; color: #f8fafc;")
        
        console_card.add_widget(self.txt_console)
        self.main_layout.addWidget(console_card, stretch=1)

    def _on_modo_toggled(self, is_checked: bool):
        """Alterna el modo de trabajo del dashboard."""
        if self.active_worker and self.active_worker.isRunning():
            self.switch_modo.setChecked(not is_checked)
            QMessageBox.warning(self, "Acción Bloqueada", "No puede cambiar de modo mientras el Bot está activo.")
            return

        if is_checked:
            self.lbl_titulo.setText("🧾 R2F-CANCÚN — MODO: FACTURACIÓN [PROXIMAMENTE]")
            self.btn_importar_excel.setEnabled(False)
            self.btn_iniciar.setEnabled(False)  # Bloqueado hasta integrar portal de facturas
            self._write_log("Modo cambiado a FACTURACIÓN. Portal de facturación pendiente de implementar.")
        else:
            self.lbl_titulo.setText("📥 R2F-CANCÚN — MODO: RECIBOS")
            self.btn_importar_excel.setEnabled(True)
            self.btn_iniciar.setEnabled(True)
            self._write_log("Modo cambiado a RECIBOS.")

    def _on_iniciar_bot(self):
        """Lanza el worker de Playwright para descargas de recibos."""
        if self.active_worker and self.active_worker.isRunning():
            return

        if not self.selected_lote_id:
            QMessageBox.warning(self, "Lote no seleccionado", "Selecciona un lote haciendo doble clic en la tabla de lotes.")
            return

        self.btn_iniciar.setEnabled(False)
        self.btn_detener.setEnabled(True)
        self.switch_modo.setEnabled(False)
        self.btn_browse.setEnabled(False)

        # Si el interruptor de visible está activo (Checked), headless debe ser False (Navegador visible)
        es_visible = self.chk_autonomo.isChecked()
        headless_mode = not es_visible

        # Inicializar el QThread Worker pasando la ruta personalizada si existe
        self.active_worker = BotReciboCunWorker(
            db_connector=self.db_connector,
            lote_id=self.selected_lote_id,
            headless=headless_mode,
            custom_output_dir=self.selected_custom_path
        )

        self.active_worker.status_changed.connect(self._write_log)
        self.active_worker.metric_updated.connect(self._on_metric_updated)
        self.active_worker.finished_processing.connect(self._on_worker_finished)
        self.active_worker.start()

        self.lbl_portal_status.setText("Portal: ACTIVO")

    def _on_detener_bot(self):
        if self.active_worker:
            self.active_worker.stop()
            self.btn_detener.setEnabled(False)

    def _on_worker_finished(self, success: bool, message: str):
        self.btn_iniciar.setEnabled(True)
        self.btn_detener.setEnabled(False)
        self.switch_modo.setEnabled(True)
        self.btn_browse.setEnabled(True)
        self.lbl_portal_status.setText("Portal: INACTIVO")

        if success:
            QMessageBox.information(self, "Proceso Finalizado", message)
        else:
            QMessageBox.critical(self, "Error del Bot", message)

        self._refresh_lotes_table()
        if self.selected_lote_id:
            self._load_lote_detalles(self.selected_lote_id)

    def _on_metric_updated(self, metric: str, value: int):
        if metric == "exitosos":
            self.box_exitosos.set_value(str(value))
        elif metric == "errores":
            self.box_errores.set_value(str(value))

    def _refresh_lotes_table(self):
        """Carga la lista de lotes desde la base de datos usando el estándar populate_rows de StyledDataTable."""
        try:
            with self.db_connector.get_session() as session:
                repo = LoteFolioRepository(session)
                lotes = repo.list_all()

                data = []
                for lote in lotes:
                    data.append([
                        str(lote.lote_id),
                        lote.folio_lote,
                        lote.origen,
                        str(lote.total_folios),
                        str(lote.folios_procesados),
                        lote.estado.codigo
                    ])
                self.table_lotes.populate_rows(data)
        except Exception as e:
            logger.error(f"Error cargando tabla de lotes: {e}")

    def _on_lote_double_clicked(self, index):
        row = index.row()
        lote_id_item = self.table_lotes.item(row, 0)
        if lote_id_item:
            lote_id = int(lote_id_item.text())
            self.selected_lote_id = lote_id
            self.lbl_lote_actual_info.setText(f"Lote seleccionado: ID {lote_id}")
            self._load_lote_detalles(lote_id)

    def _load_lote_detalles(self, lote_id: int):
        try:
            with self.db_connector.get_session() as session:
                lote_repo = LoteFolioRepository(session)
                lote = lote_repo.get_by_id(lote_id)
                if lote:
                    self.box_pendientes.set_value(str(lote.total_folios - lote.folios_procesados - lote.folios_error))
                    self.box_exitosos.set_value(str(lote.folios_procesados))
                    self.box_errores.set_value(str(lote.folios_error))

                    data = []
                    for folio in lote.folios:
                        f_text = folio.folio_electronico if folio.tipo_folio == "ELECTRONICO" else folio.folio_pase_caja
                        data.append([
                            str(folio.folio_id),
                            f_text,
                            folio.tipo_folio,
                            str(folio.intentos),
                            folio.estado.codigo
                        ])
                    self.table_detalles.populate_rows(data)
        except Exception as e:
            logger.error(f"Error cargando folios de lote: {e}")

    def _on_importar_excel(self):
        """Diálogo para seleccionar e importar un archivo Excel validando duplicados contra la base de datos."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Importar Lote Excel", "", "Archivos de Excel (*.xlsx *.xls)"
        )
        if not file_path:
            return

        try:
            from cancunbot.src.services.excel_importer import ExcelImporter
            importer = ExcelImporter()

            # Validar y cargar folios desde el archivo usando la lógica del servicio
            lista_folios_dict = importer.importar(file_path)

            if not lista_folios_dict:
                QMessageBox.warning(self, "Archivo Vacío", "No se encontraron folios validos en el archivo de excel")
                return

            folios_mapeados = []
            folios_texto_nuevos = set()
            
            # Contadores de control
            duplicados_excel = 0
            duplicados_db = 0
            total_leidos = len(lista_folios_dict)
            
            # Obtener folios ya existentes en BD para omitir duplicados
            with self.db_connector.get_session() as session:
                from sqlalchemy import select
                from cancunbot.src.storage.cancunbot_models import FolioCancun
                
                db_folios_elec = set(session.scalars(select(FolioCancun.folio_electronico).where(FolioCancun.folio_electronico.isnot(None))).all())
                db_folios_pase = set(session.scalars(select(FolioCancun.folio_pase_caja).where(FolioCancun.folio_pase_caja.isnot(None))).all())

            for f in lista_folios_dict:
                tipo = f["tipo_folio"]
                val = f["folio_electronico"] if tipo == "ELECTRONICO" else f["folio_pase_caja"]
                
                # 1. Validar duplicados dentro del propio archivo Excel
                if val in folios_texto_nuevos:
                    duplicados_excel += 1
                    continue

                # 2. Validar duplicados contra la Base de Datos
                if tipo == "ELECTRONICO" and val in db_folios_elec:
                    duplicados_db += 1
                    continue
                if tipo == "PASE_CAJA" and val in db_folios_pase:
                    duplicados_db += 1
                    continue

                folios_texto_nuevos.add(val)
                folios_mapeados.append((val, tipo))

            # Si no hay folios nuevos válidos que procesar
            if not folios_mapeados:
                msg_error = (
                    f"No hay folios nuevos para importar.\n\n"
                    f"• Total en Excel: {total_leidos}\n"
                    f"• Duplicados en Excel: {duplicados_excel}\n"
                    f"• Ya existentes en BD: {duplicados_db}"
                )
                QMessageBox.warning(self, "Importación Cancelada", msg_error)
                return

            # 3. Cuadro de Confirmación detallado antes de insertar en BD
            confirm_msg = (
                f"Resumen del archivo Excel:\n\n"
                f"• Total de registros leídos: {total_leidos}\n"
                f"• Duplicados dentro del Excel (se omitirán): {duplicados_excel}\n"
                f"• Ya registrados en Base de Datos (se omitirán): {duplicados_db}\n"
                f"• Folios nuevos a importar: {len(folios_mapeados)}\n\n"
                f"¿Deseas confirmar la inserción y crear un nuevo Lote en la Base de Datos?"
            )
            
            btn_reply = QMessageBox.question(
                self, "Confirmar Inserción", confirm_msg,
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
            )
            
            if btn_reply != QMessageBox.Yes:
                self._write_log("Importación cancelada por el usuario.")
                return

            # Proceder con la inserción física
            with self.db_connector.get_session() as session:
                lote_repo = LoteFolioRepository(session)
                folio_repo = FolioCancunRepository(session)

                lote = lote_repo.create(
                    usuario_id=self.usuario_id,
                    origen="EXCEL",
                    descripcion=f"Importado desde {Path(file_path).name}",
                    archivo_excel=file_path
                )
                
                guardados = folio_repo.create_bulk(lote.lote_id, folios_mapeados)
                lote_repo.update_metrics_and_status(lote.lote_id)
                session.commit()
                
                # Extraer atributos perezosos (lazy) del modelo ORM antes de cerrar la sesión
                folio_lote_nombre = lote.folio_lote

            QMessageBox.information(
                self, "Importación Completada", 
                f"Lote {folio_lote_nombre} creado con éxito.\n"
                f"Se insertaron {guardados} folios nuevos."
            )
            self._refresh_lotes_table()

        except Exception as err:
            logger.error(f"Error importando lote desde Excel: {err}")
            QMessageBox.critical(self, "Error de Importación", f"No se pudo procesar el archivo Excel: {err}")

    def _on_descargar_plantilla(self):
        """Genera y descarga una plantilla Excel vacía con el formato de columnas aceptado por el validador."""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla Excel", "Plantilla_Folios_Cancun.xlsx", "Archivos de Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Folios a Procesar"

            # Escribir encabezados oficiales requeridos por el importador
            ws.cell(row=1, column=1, value="FOLIO_ELECTRONICO")
            ws.cell(row=1, column=2, value="FOLIO_PASE_CAJA")

            # Ejemplo visual en la fila 2
            ws.cell(row=2, column=1, value="F-2026-615-31044")
            ws.cell(row=2, column=2, value="")
            ws.cell(row=3, column=1, value="")
            ws.cell(row=3, column=2, value="987654321")

            wb.save(file_path)
            QMessageBox.information(
                self, "Plantilla Descargada", 
                f"La plantilla se guardó en:\n{file_path}\n\nPuedes llenar los folios electrónicos en la columna 'FOLIO_ELECTRONICO' o los pases de caja en 'FOLIO_PASE_CAJA'."
            )
        except Exception as e:
            logger.error(f"Error generando plantilla Excel: {e}")
            QMessageBox.critical(self, "Error al guardar", f"No se pudo generar la plantilla: {e}")

    def _on_gear_clicked(self):
        """Muestra un menú contextual con las URLs de los portales activos (según el modo)."""
        menu = QMenu(self)
        menu.setObjectName("botMenu")

        url_recibos = "https://recibo.tesoreriacancun.com"
        url_facturas = "https://benitojuarez.expidefactura.com/"

        try:
            with self.db_connector.get_session() as session:
                repo = ConfigRepository(session)
                db_url_r = repo.get_parametro("CANCUN_PORTAL_RECIBO_URL")
                db_url_f = repo.get_parametro("CANCUN_PORTAL_FACTURA_URL")
                if db_url_r:
                    url_recibos = db_url_r
                if db_url_f:
                    url_facturas = db_url_f
        except Exception as e:
            logger.error(f"Error cargando URLs para menú de configuración: {e}")

        # Determinar cuál URL mostrar como activa según el switch de modo
        es_modo_facturas = self.switch_modo.isChecked()
        if es_modo_facturas:
            action = menu.addAction(f"🔗 Portal Facturación: {url_facturas}")
        else:
            action = menu.addAction(f"🔗 Portal Recibos: {url_recibos}")
        action.setEnabled(False)

        menu.exec_(QCursor.pos())

    def _on_user_clicked(self):
        """Muestra la información de perfil del usuario firmado y opción para desloguear."""
        menu = QMenu(self)
        menu.setObjectName("botMenu")

        nombre_usuario = f"Usuario ID: {self.usuario_id}"
        try:
            with self.db_connector.get_session() as session:
                from sar.src.storage.models import Usuario
                db_user = session.get(Usuario, self.usuario_id)
                if db_user and db_user.nombre:
                    nombre_usuario = db_user.nombre
        except Exception as e:
            logger.error(f"Error cargando perfil de usuario: {e}")

        menu.addAction(f"👤 Operador: {nombre_usuario}").setEnabled(False)
        menu.addSeparator()
        
        logout_action = menu.addAction("🚪 Cerrar Sesión")
        logout_action.triggered.connect(self.logout_requested.emit)
        
        menu.exec_(QCursor.pos())

    def _write_log(self, text: str):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.txt_console.append(f"[{timestamp}] {text}")

    def closeEvent(self, event):
        if self.active_worker and self.active_worker.isRunning():
            reply = QMessageBox.question(
                self, "Bot en ejecución",
                "El bot de descarga está activo. ¿Deseas detenerlo y cerrar?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.active_worker.stop()
                self.active_worker.wait()
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    def _verify_paths(self):
        """Replicando la lógica de verificación de acceso a la ruta de Face A."""
        if hasattr(self, 'status_indicator') and self.status_indicator:
            self.status_indicator.set_status("checking")
            
            from sar.src.core.access_manager import PathVerifyThread
            target_dir = self.selected_custom_path if self.selected_custom_path else self.default_output_dir
            
            self.path_verify_thread = PathVerifyThread(target_dir, self)
            self.path_verify_thread.result_ready.connect(self._on_path_verified)
            self.path_verify_thread.start()

    def _on_path_verified(self, path_str, has_access, error_message):
        if has_access:
            self.status_indicator.set_status("online", "CONECTADO")
            self._write_log(f"Ruta de almacenamiento accesible y verificada: {path_str}")
        else:
            self.status_indicator.set_status("offline", "SIN ACCESO")
            self._write_log(f"⚠️ ADVERTENCIA: La ruta de almacenamiento '{path_str}' no es accesible. Detalle: {error_message}")

    def _on_browse_path_clicked(self):
        """Diálogo de selección de directorio de almacenamiento, replicando la confirmación de Face A."""
        if self.active_worker and self.active_worker.isRunning():
            return

        reply = QMessageBox.question(
            self,
            "Confirmar Cambio de Ruta",
            f"Se recomienda utilizar la ruta por defecto ({self.default_output_dir}) para la sincronización y auditoría.\n\n¿Estás seguro de que deseas cambiar la ruta de descarga?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
            
        dir_path = QFileDialog.getExistingDirectory(self, "Seleccionar Carpeta de Almacenamiento de Recibos")
        if dir_path:
            self.selected_custom_path = dir_path
            # Truncar visualmente la ruta si es demasiado larga
            display_path = dir_path if len(dir_path) < 35 else "..." + dir_path[-32:]
            self.lbl_download_path_display.setText(display_path)
            self._write_log(f"Ruta de almacenamiento cambiada a: {dir_path}")
            self._verify_paths()
