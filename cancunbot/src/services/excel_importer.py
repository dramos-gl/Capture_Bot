"""
CancunBot — Servicio de Importación de Folios desde Excel
Lee una hoja de cálculo y retorna la lista de folios a procesar.
"""
import logging
from pathlib import Path
from typing import Optional

import openpyxl

from sqlalchemy import text
from sar.src.storage.repositories import ConfigRepository
from sar.src.storage.db_connector import DatabaseConnector

logger = logging.getLogger(__name__)


class ExcelImporter:
    """
    Importa folios desde un archivo Excel.
    
    Estructura esperada del Excel:
        - Columna 'FOLIO_ELECTRONICO' (o el nombre configurado en BD): Folio electrónico
        - Columna 'FOLIO_PASE_CAJA' (opcional): Folio de pase de caja
    
    Los nombres de columna son configurables via parámetros de sistema:
        - CANCUN_EXCEL_COL_FOLIO_ELECTRONICO
        - CANCUN_EXCEL_COL_FOLIO_PASE_CAJA
    """

    def __init__(self):
        # Conectar dinámicamente a la BD del SAR para obtener las configuraciones reales
        db = DatabaseConnector()
        with db.get_session() as session:
            repo = ConfigRepository(session)
            self._col_electronico = repo.get_parametro("CANCUN_EXCEL_COL_FOLIO_ELECTRONICO") or "FOLIO_ELECTRONICO"
            self._col_pase_caja = repo.get_parametro("CANCUN_EXCEL_COL_FOLIO_PASE_CAJA") or "FOLIO_PASE_CAJA"

    def importar(self, ruta_excel: str) -> list[dict]:
        """
        Lee el archivo Excel y retorna la lista de folios.
        
        Args:
            ruta_excel: Ruta al archivo .xlsx
        
        Returns:
            Lista de dicts con claves: folio_electronico, folio_pase_caja, tipo_folio
        
        Raises:
            FileNotFoundError: Si el archivo no existe
            ValueError: Si el archivo no tiene las columnas requeridas
        """
        ruta = Path(ruta_excel)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_excel}")

        logger.info(f"Importando folios desde: {ruta_excel}")

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active

        # Leer encabezados de la primera fila y limpiarlos (removiendo espacios, guiones y guiones bajos para máxima compatibilidad)
        def _clean_header(name: str) -> str:
            return name.strip().upper().replace(" ", "").replace("_", "").replace("-", "")

        headers_original = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_cleaned = [_clean_header(h) for h in headers_original]

        target_elec = _clean_header(self._col_electronico)
        target_pase = _clean_header(self._col_pase_caja)

        # Fallbacks amigables
        fallbacks_elec = [target_elec, "FOLIO", "FOLIOELECTRONICO", "FOLIOELEC", "ELECTRONICO"]
        fallbacks_pase = [target_pase, "PASECAJA", "FOLIOPASECAJA", "PASE", "FOLIO_PASE_CAJA"]

        idx_elec: Optional[int] = None
        idx_pase: Optional[int] = None

        # Buscar coincidencia para FOLIO_ELECTRONICO
        for fallback in fallbacks_elec:
            if fallback in headers_cleaned:
                idx_elec = headers_cleaned.index(fallback)
                break

        # Buscar coincidencia para FOLIO_PASE_CAJA
        for fallback in fallbacks_pase:
            if fallback in headers_cleaned:
                idx_pase = headers_cleaned.index(fallback)
                break

        # Buscar coincidencia para RFC y DESARROLLO
        idx_rfc: Optional[int] = None
        idx_desarrollo: Optional[int] = None
        for i, h in enumerate(headers_cleaned):
            if h in ["RFC", "RFC_EMPRESA", "CLIENTE_RFC"]:
                idx_rfc = i
            elif h in ["DESARROLLO", "NOMBRE_DESARROLLO", "PROYECTO"]:
                idx_desarrollo = i

        if idx_elec is None and idx_pase is None:
            raise ValueError(
                f"El Excel no tiene las columnas esperadas.\n"
                f"Se requiere una columna como '{self._col_electronico}' o '{self._col_pase_caja}'.\n"
                f"Columnas detectadas en el archivo: {headers_original}"
            )

        folios: list[dict] = []
        db = DatabaseConnector()
        with db.get_session() as session:
            # Caché de catálogos en memoria para velocidad
            res_rfc = session.execute(text("SELECT rfc_id, rfc FROM sar_catalogo.rfc WHERE activo = true")).fetchall()
            rfc_cache = {row[1].strip().upper(): row[0] for row in res_rfc}

            res_des = session.execute(text("SELECT desarrollo_id, nombre FROM sar_catalogo.desarrollo WHERE activo = true")).fetchall()
            def _clean(val: str) -> str:
                return val.strip().upper().replace(" ", "").replace("_", "").replace("-", "")
            desarrollo_cache = {_clean(row[1]): row[0] for row in res_des}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) == 0:
                    continue
                folio_elec = str(row[idx_elec]).strip() if idx_elec is not None and row[idx_elec] is not None else None
                folio_pase = str(row[idx_pase]).strip() if idx_pase is not None and row[idx_pase] is not None else None

                # Saltar filas vacías
                if not folio_elec and not folio_pase:
                    continue

                # Determinar tipo de folio
                tipo = "ELECTRONICO" if folio_elec else "PASE_CAJA"

                # Resolver RFC
                rfc_val = str(row[idx_rfc]).strip().upper() if idx_rfc is not None and row[idx_rfc] is not None else None
                rfc_id = rfc_cache.get(rfc_val) if rfc_val else None

                # Resolver Desarrollo
                des_val = str(row[idx_desarrollo]).strip() if idx_desarrollo is not None and row[idx_desarrollo] is not None else None
                des_id = desarrollo_cache.get(_clean(des_val)) if des_val else None

                folios.append({
                    "folio_electronico": folio_elec,
                    "folio_pase_caja": folio_pase,
                    "tipo_folio": tipo,
                    "rfc_id": rfc_id,
                    "desarrollo_id": des_id,
                    "excel_rfc": rfc_val,
                    "excel_desarrollo": des_val
                })

        wb.close()
        logger.info(f"Importados {len(folios)} folios desde el Excel con resolución de catálogos.")
        return folios

    def validar_estructura(self, ruta_excel: str) -> tuple[bool, str]:
        """
        Valida que el archivo Excel tenga la estructura correcta sin importar datos.
        
        Returns:
            Tuple (es_valido, mensaje)
        """
        try:
            ruta = Path(ruta_excel)
            if not ruta.exists():
                return False, f"Archivo no encontrado: {ruta_excel}"
            if ruta.suffix.lower() not in [".xlsx", ".xls"]:
                return False, "El archivo debe ser formato Excel (.xlsx o .xls)"
            self.importar(ruta_excel)  # Prueba de importación
            return True, "Estructura válida"
        except ValueError as e:
            return False, str(e)
        except Exception as e:
            return False, f"Error al validar: {e}"
