import os
import openpyxl
from datetime import datetime, date
from typing import List, Dict, Any
from sqlalchemy import select, and_

class ExcelInventoryHandler:
    """Service to parse and validate Excel files for reference inventory control."""

    @staticmethod
    def parse_excel_inventory(file_path: str) -> List[Dict[str, Any]]:
        """Parses the Excel file and normalizes it into a list of detail dictionaries.
        
        One row in Excel with multiple reference columns (ANALISIS, AVISO, CLG)
        will be split into multiple detail records.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"El archivo no existe: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active # Use the active sheet (typically Hoja1)
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Find header indices
        headers = [str(h).strip().upper() for h in rows[0]]
        
        def get_idx(*names):
            for name in names:
                name_upper = name.strip().upper()
                if name_upper in headers:
                    return headers.index(name_upper)
            return -1

        idx_empresa = get_idx("RAZON SOCIAL", "RAZÓN SOCIAL", "EMPRESA", "RFC", "EMPRESA FACTURADORA")
        idx_desarrollo = get_idx("DESARROLLO")
        idx_delegacion = get_idx("DELEGACION", "DELEGACIÓN")
        idx_mz = get_idx("MZA", "MZ", "MANZANA")
        idx_lote = get_idx("LOTE", "LT")
        idx_edif = get_idx("EXT", "EDIF", "EDIFICIO", "EXTERIOR")
        idx_viv = get_idx("INT", "VIV", "VIVIENDA", "INTERIOR")
        idx_folio = get_idx("No.OFICIAL", "No. OFICIAL", "NO.OFICIAL", "NO. OFICIAL", "FOLIO")
        idx_cliente = get_idx("CLIENTE", "NOMBRE CLIENTE", "NOMBRE")
        idx_estatus_aviso = get_idx("FECHA INGRESO A RPP", "FECHA_INGRESO_A_RPP", "ESTATUS_PRIMER_AVISO", "ESTATUS RPP")
        idx_credito_titular = get_idx("CREDITO_TITULAR", "CREDITO TITULAR", "CREDITO", "TITULAR", "CREDITO_O_TITULAR")
        idx_pa = get_idx("P.A", "P.A.", "PA", "PADRON", "PADRÓN")
        idx_comentarios = get_idx("COMENTARIOS", "OBSERVACIONES", "OBS")
        idx_fecha_sol = get_idx("FECHA_SOLICITUD", "FECHA SOLICITUD", "FECHA")
        idx_ubicacion = get_idx("UBICACION", "UBICACIÓN")
        idx_fecha_reporte_notaria = get_idx("FECHA REPORTA LA NOTARIA", "FECHA_REPORTA_LA_NOTARIA", "FECHA REPORTA NOTARIA")
        idx_fecha_escritura = get_idx("FECHA ESCRITURA", "FECHA_ESCRITURA")
        idx_fecha_titulacion = get_idx("FECHA DE TITULACION", "FECHA DE TITULACIÓN", "FECHA TITULACION", "FECHA_TITULACION")
        
        # Concept reference columns (resilient to accents, spaces, and abbreviations)
        concept_cols = {
            "ANALISIS": get_idx("ANALISIS", "ANÁLISIS", "ANALIS"),
            "AVISO": get_idx("AVISO", "1ER_AVISO", "1ER _AVISO", "AVISO_RPP", "PRIMER AVISO", "AVISO PREVENTIVO"),
            "CLG": get_idx("CLG", "CERTIFICADO", "CLG_RPP"),
            "CANC_1ER _AVISO": get_idx("CANCELACION PRIMER AVISO", "CANCELACIÓN PRIMER AVISO", "CANC_1ER _AVISO", "CANC_1ER_AVISO", "CANCELACION_1ER_AVISO", "CANCELACIÓN 1ER AVISO", "CANCELACION 1ER AVISO"),
            "CANC_2DO_AVISO": get_idx("CANCELACION SEGUNDO AVISO", "CANCELACIÓN SEGUNDO AVISO", "CANC_2DO_AVISO", "CANCELACION_2DO_AVISO", "CANCELACIÓN 2DO AVISO", "CANCELACION 2DO AVISO"),
            "NUEVO_DERECHO_AVISO": get_idx("NUEVO_DERECHO_AVISO", "NUEVO_AVISO", "NUEVO DERECHO AVISO"),
        }

        parsed_records = []

        # Process data rows
        for row_num, row in enumerate(rows[1:], start=2):
            # Skip empty rows
            if not row or all(v is None for v in row):
                continue
                
            cliente = str(row[idx_cliente]).strip() if idx_cliente != -1 and row[idx_cliente] is not None else ""
            desarrollo = str(row[idx_desarrollo]).strip().upper() if idx_desarrollo != -1 and row[idx_desarrollo] is not None else ""
            empresa = str(row[idx_empresa]).strip().upper() if idx_empresa != -1 and row[idx_empresa] is not None else ""

            # Check if row has at least one concept reference
            has_reference = False
            for col_idx in concept_cols.values():
                if col_idx != -1 and row[col_idx] is not None and str(row[col_idx]).strip() != "":
                    has_reference = True
                    break

            if not cliente and not has_reference:
                continue # Skip rows if neither client nor references exist
            
            # Format date
            fecha_sol = None
            if idx_fecha_sol != -1 and row[idx_fecha_sol] is not None:
                val = row[idx_fecha_sol]
                if isinstance(val, (datetime, date)):
                    fecha_sol = val
                else:
                    try:
                        fecha_sol = datetime.strptime(str(val).split()[0], "%Y-%m-%d").date()
                    except Exception:
                        pass
            
            # Location fields
            mz = str(row[idx_mz]).strip() if idx_mz != -1 and row[idx_mz] is not None else ""
            lote = str(row[idx_lote]).strip() if idx_lote != -1 and row[idx_lote] is not None else ""
            edif = str(row[idx_edif]).strip() if idx_edif != -1 and row[idx_edif] is not None else ""
            viv = str(row[idx_viv]).strip() if idx_viv != -1 and row[idx_viv] is not None else ""
            
            raw_ubicacion = str(row[idx_ubicacion]).strip() if idx_ubicacion != -1 and row[idx_ubicacion] is not None else ""

            # Parse fallback from UBICACION text column if mz/lote are empty
            if (not mz and not lote) and idx_ubicacion != -1 and row[idx_ubicacion] is not None:
                # Try parsing e.g. "MZ 38 LT 2 VIV 35"
                import re
                txt = raw_ubicacion.upper()
                m_mz = re.search(r"MZ\s*(\d+)", txt)
                m_lt = re.search(r"L?T\s*(\d+)", txt)
                m_ed = re.search(r"EDIF\s*([A-Z0-9\-]+)", txt)
                m_vv = re.search(r"VIV\s*([A-Z0-9\-]+)", txt)
                if m_mz: mz = m_mz.group(1)
                if m_lt: lote = m_lt.group(1)
                if m_ed: edif = m_ed.group(1)
                if m_vv: viv = m_vv.group(1)

            folio = str(row[idx_folio]).strip() if idx_folio != -1 and row[idx_folio] is not None else ""
            estatus_aviso = str(row[idx_estatus_aviso]).strip() if idx_estatus_aviso != -1 and row[idx_estatus_aviso] is not None else ""
            credito_titular = str(row[idx_credito_titular]).strip() if idx_credito_titular != -1 and row[idx_credito_titular] is not None else ""
            pa = str(row[idx_pa]).strip() if idx_pa != -1 and row[idx_pa] is not None else ""
            comentarios = str(row[idx_comentarios]).strip() if idx_comentarios != -1 and row[idx_comentarios] is not None else ""
            delegacion = str(row[idx_delegacion]).strip() if idx_delegacion != -1 and row[idx_delegacion] is not None else ""
            
            f_rep_notaria = row[idx_fecha_reporte_notaria] if idx_fecha_reporte_notaria != -1 else None
            f_escritura = row[idx_fecha_escritura] if idx_fecha_escritura != -1 else None
            f_titulacion = row[idx_fecha_titulacion] if idx_fecha_titulacion != -1 else None

            # Check each concept reference column
            for concept_name, col_idx in concept_cols.items():
                if col_idx != -1 and row[col_idx] is not None:
                    ref_val = str(row[col_idx]).strip()
                    # Skip empty/None references or headers accidentally repeated
                    if not ref_val or ref_val.upper() in ("NONE", "NULL", "-", ""):
                        continue
                    
                    is_indicator = False
                    if len(ref_val) <= 4 or ref_val.upper() in ("X", "SI", "S", "1", "YES", "OK", "X/"):
                        is_indicator = True
                    
                    # Store record
                    parsed_records.append({
                        "excel_row": row_num,
                        "cliente": cliente,
                        "desarrollo": desarrollo,
                        "empresa": empresa,
                        "fecha_solicitud": fecha_sol,
                        "ubicacion": raw_ubicacion,
                        "mz": mz,
                        "lote": lote,
                        "edif": edif,
                        "viv": viv,
                        "folio_electronico": folio,
                        "estatus_primer_aviso": estatus_aviso,
                        "credito_titular": credito_titular,
                        "pa": pa,
                        "comentarios": comentarios,
                        "fecha_reporte_notaria": f_rep_notaria,
                        "fecha_escritura": f_escritura,
                        "fecha_titulacion": f_titulacion,
                        "delegacion": delegacion,
                        "concepto_solicitado": concept_name,
                        "referencia_asignada": "" if is_indicator else ref_val,
                        "requiere_autovincular": is_indicator
                    })

        return parsed_records

    @staticmethod
    def validate_parsed_rows(
        session, parsed_rows: List[Dict[str, Any]], default_rfc_id: Optional[int] = None, completar_notaria_id: Optional[int] = None, orden_ids: Optional[List[int]] = None, solo_reservar: bool = False
    ) -> List[Dict[str, Any]]:
        """Validates parsed rows against the database, enforcing:
        1. Reference exists and is FACTURADA (or auto-assigns an available one if empty).
        2. Reference is not already assigned.
        3. Concept matching (CLG, AVISO, etc.).
        4. Geolocation match (desarrollo delegation == reference delegation).
        5. Company (RFC) matching for autolink and validation.
        6. If completing a reservation (completar_notaria_id), matches against RESERVADA placeholders using FIFO.
        """
        from sar.src.storage.models import Referencia, EstadoSistema, Desarrollo, Delegacion, GrupoReferencia, Concepto, Solicitud, Rfc, Ubicacion, AsignacionReferencia, LoteAsignacion, LoteDetalle
        from sqlalchemy import select
        
        # Cache concepts mapping
        concepto_stmt = select(Concepto)
        concepts = session.execute(concepto_stmt).scalars().all()
        concept_alias_map = {c.alias: c for c in concepts if c.alias}

        # Cache RFCs mapping (razon_social and rfc strings to id)
        rfcs_stmt = select(Rfc).where(Rfc.activo == True)
        rfcs = session.execute(rfcs_stmt).scalars().all()
        rfcs_map = {r.razon_social.strip().upper(): r.rfc_id for r in rfcs}
        rfcs_map.update({r.rfc.strip().upper(): r.rfc_id for r in rfcs})
        
        # Load active RESERVADA placeholders if completing a notary reservation
        reserved_placeholders = []
        if completar_notaria_id:
            placeholder_stmt = (
                select(AsignacionReferencia, Referencia)
                .join(LoteDetalle, AsignacionReferencia.lote_detalle_id == LoteDetalle.lote_detalle_id)
                .join(LoteAsignacion, LoteDetalle.lote_asignacion_id == LoteAsignacion.lote_asignacion_id)
                .join(Referencia, AsignacionReferencia.referencia_id == Referencia.referencia_id)
                .join(EstadoSistema, Referencia.estado_id == EstadoSistema.estado_id)
                .where(
                    LoteAsignacion.notaria_id == completar_notaria_id,
                    EstadoSistema.entidad == 'referencia',
                    EstadoSistema.codigo == 'RESERVADA',
                    AsignacionReferencia.ubicacion_id.is_(None)
                )
                .order_by(LoteAsignacion.fecha.asc(), AsignacionReferencia.asignacion_referencia_id.asc())
            )
            reserved_placeholders = session.execute(placeholder_stmt).all()
            # Convert to a mutable list of dicts/tuples to pop sequentially
            reserved_placeholders = list(reserved_placeholders)

        validated_rows = []
        allocated_ref_ids = set()

        for row in parsed_rows:
            ref_str = row["referencia_asignada"]
            concept_req = row["concepto_solicitado"]
            desarrollo_name = row["desarrollo"]
            row_empresa = row.get("empresa", "").strip().upper()
            req_autolink = row.get("requiere_autovincular", False)
            cliente = row.get("cliente", "")
            mz = row.get("mz", "")
            lote = row.get("lote", "")
            edif = row.get("edif", "")
            viv = row.get("viv", "")
            
            row_result = dict(row)
            row_result["status"] = "PENDIENTE" # default
            row_result["error_message"] = ""
            row_result["referencia_id"] = None
            row_result["desarrollo_id"] = None
            row_result["delegacion_nombre"] = ""
            row_result["rfc_id"] = None
            row_result["lote_detalle_id"] = None  # To track which reservation to update

            # Validate critical missing fields presence: cliente, mz, lote, ext (edif), int (viv)
            missing_fields = []
            if not solo_reservar:
                if not cliente or cliente.strip() == "": missing_fields.append("Cliente")
                if not mz or mz.strip() == "": missing_fields.append("MZA")
                if not lote or lote.strip() == "": missing_fields.append("Lote")
                if not edif or edif.strip() == "": missing_fields.append("Ext (Exterior)")
                if not viv or viv.strip() == "": missing_fields.append("Int (Interior)")
            
            if missing_fields:
                row_result["status"] = "ERROR"
                row_result["error_message"] = f"Faltan campos obligatorios de la ubicación: {', '.join(missing_fields)}."
                validated_rows.append(row_result)
                continue
            
            # Resolve Company (RFC)
            resolved_rfc_id = None
            if row_empresa:
                resolved_rfc_id = rfcs_map.get(row_empresa)
                if not resolved_rfc_id:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = f"La empresa '{row_empresa}' no está registrada o no está activa."
                    validated_rows.append(row_result)
                    continue
            else:
                resolved_rfc_id = default_rfc_id

            row_result["rfc_id"] = resolved_rfc_id
            
            # 1. Lookup/register Desarrollo first
            if solo_reservar and not desarrollo_name:
                desarrollo = None
                row_result["desarrollo_id"] = None
                deleg_name = "SIN ASIGNAR"
                row_result["delegacion_nombre"] = deleg_name
            else:
                if not desarrollo_name:
                    desarrollo_name = "GENERICO"
                des_stmt = select(Desarrollo).where(Desarrollo.nombre == desarrollo_name)
                desarrollo = session.execute(des_stmt).scalars().first()
                
                if not desarrollo:
                    desarrollo = Desarrollo(nombre=desarrollo_name, activo=True)
                    session.add(desarrollo)
                    session.flush()

                    # Default delegation logic
                    deleg_id = 2 # default Cancun
                    if "PLAYA" in desarrollo_name or "TULUM" in desarrollo_name:
                        deleg_id = 3 # Playa del Carmen

                    from sar.src.storage.models import DesarrolloEmpresa
                    # resolved_rfc_id fallback to 1 if empty
                    de = DesarrolloEmpresa(
                        desarrollo_id=desarrollo.desarrollo_id,
                        rfc_id=resolved_rfc_id if resolved_rfc_id else 1,
                        delegacion_id=deleg_id,
                        es_default=True,
                        activo=True
                    )
                    session.add(de)
                    session.flush()

                row_result["desarrollo_id"] = desarrollo.desarrollo_id

                # Fetch delegation_id from DesarrolloEmpresa associations
                from sar.src.storage.models import DesarrolloEmpresa
                de_stmt = select(DesarrolloEmpresa.delegacion_id).where(DesarrolloEmpresa.desarrollo_id == desarrollo.desarrollo_id)
                deleg_id = session.execute(de_stmt).scalars().first()
                if not deleg_id:
                    deleg_id = 2 # default Cancun fallback

                # Fetch delegation name details
                deleg_stmt = select(Delegacion.nombre).where(Delegacion.delegacion_id == deleg_id)
                deleg_name = session.execute(deleg_stmt).scalar()
                row_result["delegacion_nombre"] = deleg_name

            # Check if this client already has an assignment for the same concept at this exact location
            has_dup = False
            dup_ref = None
            if not solo_reservar:
                normalized_concept_req = concept_req
                if normalized_concept_req == "AVISO":
                    normalized_concept_req = "AVISO PREVENTIVO"

                dup_stmt = (
                    select(AsignacionReferencia)
                    .join(LoteDetalle, AsignacionReferencia.lote_detalle_id == LoteDetalle.lote_detalle_id)
                    .join(Ubicacion, AsignacionReferencia.ubicacion_id == Ubicacion.ubicacion_id)
                    .join(Concepto, LoteDetalle.concepto_id == Concepto.concepto_id)
                    .where(
                        AsignacionReferencia.cliente == cliente,
                        Ubicacion.desarrollo_id == (desarrollo.desarrollo_id if desarrollo else None),
                        Ubicacion.mz == mz,
                        Ubicacion.lote == lote,
                        Ubicacion.edif == edif,
                        Ubicacion.viv == viv,
                        Concepto.alias == normalized_concept_req
                    )
                )
                dup_check = session.execute(dup_stmt).scalars().first()
                has_dup = dup_check is not None
                dup_ref = dup_check.referencia.referencia_portal if dup_check else None

            # Scenario A: We are completing a Notary Reservation
            if completar_notaria_id:
                # Find matching placeholder by concept
                placeholder_match = None
                placeholder_idx = -1
                
                concept_obj = concept_alias_map.get(normalized_concept_req)
                concept_id_req = concept_obj.concepto_id if concept_obj else None
                
                if req_autolink or not ref_str:
                    # Find first placeholder matching concept
                    for idx, (ld_p, ref_p) in enumerate(reserved_placeholders):
                        if ld_p.lote_detalle.concepto_id == concept_id_req:
                            placeholder_match = (ld_p, ref_p)
                            placeholder_idx = idx
                            break
                else:
                    # Find specific placeholder matching reference string
                    for idx, (ld_p, ref_p) in enumerate(reserved_placeholders):
                        if ref_p.referencia_portal == ref_str and ld_p.lote_detalle.concepto_id == concept_id_req:
                            placeholder_match = (ld_p, ref_p)
                            placeholder_idx = idx
                            break

                if not placeholder_match:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = f"No hay facturas RESERVADAS disponibles para el concepto '{concept_req}' (DB: '{normalized_concept_req}') de esta Notaría."
                    validated_rows.append(row_result)
                    continue

                ld_p, ref_p = placeholder_match
                # Pop from cached list so it is not double-allocated
                reserved_placeholders.pop(placeholder_idx)

                row_result["referencia_id"] = ref_p.referencia_id
                row_result["referencia_asignada"] = ref_p.referencia_portal
                row_result["lote_detalle_id"] = ld_p.asignacion_referencia_id

                # Warning if development doesn't match the reserved one
                if ld_p.lote_detalle.desarrollo_id != desarrollo.desarrollo_id:
                    row_result["status"] = "WARNING"
                    row_result["error_message"] = f"El desarrollo no coincide con el desarrollo reservado en el lote original ({ld_p.lote_detalle.lote_asignacion_id})."
                elif has_dup:
                    row_result["status"] = "WARNING"
                    row_result["error_message"] = f"El cliente ya tiene una asignación de {concept_req} en esta ubicación (Ref: {dup_ref})."
                else:
                    row_result["status"] = "CORRECTO"

                validated_rows.append(row_result)
                continue

            # Scenario B: General bulk assignment (Reactivo)
            if req_autolink or not ref_str:
                if not resolved_rfc_id:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = "Debe especificar la Empresa en el Excel o seleccionar una Empresa por Defecto en la pantalla."
                    validated_rows.append(row_result)
                    continue

                expected_aliases = []
                if concept_req == "CLG": expected_aliases = ["CLG"]
                elif concept_req in ("AVISO", "NUEVO_DERECHO_AVISO"): expected_aliases = ["AVISO PREVENTIVO"]
                elif concept_req == "ANALISIS": expected_aliases = ["ANALISIS"]
                else: expected_aliases = [concept_req]

                # Find a reference in FACTURADA state for this concept, delegation, and company
                available_stmt = (
                    select(Referencia)
                    .join(EstadoSistema, Referencia.estado_id == EstadoSistema.estado_id)
                    .join(GrupoReferencia, Referencia.grupo_id == GrupoReferencia.grupo_id)
                    .join(Concepto, GrupoReferencia.concepto_id == Concepto.concepto_id)
                    .join(Solicitud, Referencia.solicitud_id == Solicitud.solicitud_id)
                    .where(
                        EstadoSistema.entidad == 'referencia',
                        EstadoSistema.codigo == 'FACTURADA',
                        Concepto.alias.in_(expected_aliases),
                        Solicitud.delegacion_id == desarrollo.delegacion_id,
                        GrupoReferencia.rfc_id == resolved_rfc_id
                    )
                )
                if orden_ids:
                    available_stmt = available_stmt.where(GrupoReferencia.orden_id.in_(orden_ids))

                if allocated_ref_ids:
                    available_stmt = available_stmt.where(Referencia.referencia_id.not_in(list(allocated_ref_ids)))

                ref_obj = session.execute(available_stmt).scalars().first()
                if not ref_obj:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = f"No hay facturas disponibles ('FACTURADA') para '{concept_req}' de la empresa seleccionada en '{deleg_name}'."
                    validated_rows.append(row_result)
                    continue

                row_result["referencia_id"] = ref_obj.referencia_id
                row_result["referencia_asignada"] = ref_obj.referencia_portal
                allocated_ref_ids.add(ref_obj.referencia_id)
                if has_dup:
                    row_result["status"] = "WARNING"
                    row_result["error_message"] = f"El cliente ya tiene una asignación de {concept_req} para esta ubicación en un lote anterior (Ref: {dup_ref})."
                else:
                    row_result["status"] = "CORRECTO"
                validated_rows.append(row_result)
                continue

            # 2. Query reference details
            ref_stmt = (
                select(Referencia, EstadoSistema.codigo, Concepto.alias, Delegacion.nombre, Delegacion.delegacion_id, GrupoReferencia.orden_id)
                .join(EstadoSistema, Referencia.estado_id == EstadoSistema.estado_id)
                .join(GrupoReferencia, Referencia.grupo_id == GrupoReferencia.grupo_id)
                .join(Concepto, GrupoReferencia.concepto_id == Concepto.concepto_id)
                .join(Solicitud, Referencia.solicitud_id == Solicitud.solicitud_id)
                .join(Delegacion, Solicitud.delegacion_id == Delegacion.delegacion_id)
                .where(Referencia.referencia_portal == ref_str)
            )
            ref_res = session.execute(ref_stmt).first()

            if not ref_res:
                row_result["status"] = "ERROR"
                row_result["error_message"] = f"La referencia '{ref_str}' no existe en el sistema."
                validated_rows.append(row_result)
                continue
                
            ref_obj, estado_cod, concept_alias, ref_deleg_name, ref_deleg_id, ref_orden_id = ref_res
            
            # Check if reference belongs to the selected orders
            if orden_ids and ref_orden_id not in orden_ids:
                row_result["status"] = "ERROR"
                row_result["error_message"] = f"La referencia '{ref_str}' pertenece a una orden que no está seleccionada en los filtros."
                validated_rows.append(row_result)
                continue
                
            row_result["referencia_id"] = ref_obj.referencia_id
            row_result["concepto_solicitado"] = concept_alias

            # 3. Check if reference is in FACTURADA state
            if estado_cod != "FACTURADA":
                # Query AsignacionReferencia and join with Ubicacion to resolve the correct assigned client and format the message
                from sar.src.storage.models import AsignacionReferencia, Ubicacion
                ar_check = session.execute(
                    select(AsignacionReferencia)
                    .join(Ubicacion, AsignacionReferencia.ubicacion_id == Ubicacion.ubicacion_id)
                    .where(AsignacionReferencia.referencia_id == ref_obj.referencia_id)
                ).scalars().first()
                if ar_check and ar_check.ubicacion:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = f"La referencia ya está asignada al cliente '{ar_check.ubicacion.cliente}'."
                else:
                    row_result["status"] = "ERROR"
                    row_result["error_message"] = f"La referencia ya está asignada en el sistema (Estado: '{estado_cod}')."
                validated_rows.append(row_result)
                continue

            # 4. Check Concept Match
            # Maps column names to DB concept aliases: e.g. column 'AVISO' matches alias 'AVISO PREVENTIVO'
            expected_aliases = []
            if concept_req == "CLG": expected_aliases = ["CLG"]
            elif concept_req in ("AVISO", "NUEVO_DERECHO_AVISO"): expected_aliases = ["AVISO PREVENTIVO"]
            elif concept_req == "ANALISIS": expected_aliases = ["ANALISIS"]

            if not solo_reservar and expected_aliases and concept_alias not in expected_aliases:
                row_result["status"] = "ERROR"
                row_result["error_message"] = f"Concepto incorrecto: Referencia es de tipo '{concept_alias}' pero se solicitó '{concept_req}'."
                validated_rows.append(row_result)
                continue

            # 5. Check Geolocation Match
            if not solo_reservar and ref_deleg_id != deleg_id:
                row_result["status"] = "ERROR"
                row_result["error_message"] = f"Error geográfico: Referencia pertenece a '{ref_deleg_name}' pero el desarrollo '{desarrollo_name}' pertenece a '{deleg_name}'."
                validated_rows.append(row_result)
                continue

            # Valid reference!
            if has_dup:
                row_result["status"] = "WARNING"
                row_result["error_message"] = f"El cliente ya tiene una asignación de {concept_req} para esta ubicación en un lote anterior (Ref: {dup_ref})."
            else:
                row_result["status"] = "CORRECTO"
            validated_rows.append(row_result)

        return validated_rows

    @staticmethod
    def generate_excel_inventory_file(dest_path: str, title: str, subtitle: str, date_range: str, data_rows: List[Dict[str, Any]]):
        """Generates a styled Excel sheet mapping the structure shown in Image 1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        # Styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        font_title = Font(name="Calibri", size=16, bold=True)
        font_subtitle = Font(name="Calibri", size=12, bold=True)
        font_header = Font(name="Calibri", size=10, bold=True, color="000000")
        font_data = Font(name="Calibri", size=10)
        
        fill_header = PatternFill(start_color="86C232", end_color="86C232", fill_type="solid") # Sleek green
        
        border_thin = Side(border_style="thin", color="D3D3D3")
        border_double = Side(border_style="double", color="000000")
        border_data = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        # 1. Header block
        ws.merge_cells("A2:P2")
        ws["A2"] = title.upper()
        ws["A2"].font = font_title
        ws["A2"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A3:P3")
        ws["A3"] = subtitle.upper()
        ws["A3"].font = font_subtitle
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:P4")
        ws["A4"] = f"FECHA CORTE: {date_range.upper()}"
        ws["A4"].font = Font(name="Calibri", size=11, bold=True, italic=True)
        ws["A4"].alignment = Alignment(horizontal="center")

        # Column headers matching Control_Inventario.xlsx format
        headers = [
            "CLIENTE", "DESARROLLO", "FECHA_SOLICITUD", "UBICACION", "MZ", "LOTE", "EDIF", "VIV", 
            "FOLIO", "ESTATUS_PRIMER_AVISO", "ANALISIS", "AVISO", "CLG", "CANC_1ER _AVISO", "CANC_2DO_AVISO", "NUEVO_DERECHO_AVISO"
        ]
        
        ws.row_dimensions[6].height = 25
        for col_idx, text_header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_idx, value=text_header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=Side(border_style="medium", color="000000"))

        # We need to pivot data_rows (which are normalized in details) back to a single row per client
        # to generate a clean sheet similar to Control_Inventario.xlsx
        pivot_data = {}
        for r in data_rows:
            # Group key by client and location coordinates
            key = (r["cliente"], r["desarrollo"], r["mz"], r["lote"], r["edif"], r["viv"], r["folio_electronico"])
            if key not in pivot_data:
                pivot_data[key] = {
                    "cliente": r["cliente"],
                    "desarrollo": r["desarrollo"],
                    "fecha_solicitud": r.get("fecha_solicitud", ""),
                    "mz": r["mz"],
                    "lote": r["lote"],
                    "edif": r["edif"],
                    "viv": r["viv"],
                    "folio": r["folio_electronico"],
                    "estatus": r.get("estatus_primer_aviso", ""),
                    "references": {}
                }
            concept = r["concepto"]
            pivot_data[key]["references"][concept] = r["referencia"]

        row_num = 7
        for key, p in pivot_data.items():
            ws.row_dimensions[row_num].height = 18
            
            # Build location string
            loc_str = ""
            if p["mz"]: loc_str += f"MZ {p['mz']} "
            if p["lote"]: loc_str += f"LT {p['lote']} "
            if p["edif"]: loc_str += f"EDIF {p['edif']} "
            if p["viv"]: loc_str += f"VIV {p['viv']} "
            loc_str = loc_str.strip()

            ws.cell(row=row_num, column=1, value=p["cliente"])
            ws.cell(row=row_num, column=2, value=p["desarrollo"])
            ws.cell(row=row_num, column=3, value=p["fecha_solicitud"])
            ws.cell(row=row_num, column=4, value=loc_str)
            ws.cell(row=row_num, column=5, value=p["mz"])
            ws.cell(row=row_num, column=6, value=p["lote"])
            ws.cell(row=row_num, column=7, value=p["edif"] or None)
            ws.cell(row=row_num, column=8, value=p["viv"])
            ws.cell(row=row_num, column=9, value=p["folio"])
            ws.cell(row=row_num, column=10, value=p["estatus"])
            
            # Concept references columns mapping
            ws.cell(row=row_num, column=11, value=p["references"].get("ANALISIS", ""))
            ws.cell(row=row_num, column=12, value=p["references"].get("AVISO", ""))
            ws.cell(row=row_num, column=13, value=p["references"].get("CLG", ""))
            ws.cell(row=row_num, column=14, value=p["references"].get("CANC_1ER _AVISO", ""))
            ws.cell(row=row_num, column=15, value=p["references"].get("CANC_2DO_AVISO", ""))
            ws.cell(row=row_num, column=16, value=p["references"].get("NUEVO_DERECHO_AVISO", ""))
            
            for c in range(1, 17):
                cell = ws.cell(row=row_num, column=c)
                cell.font = font_data
                cell.border = border_data
                if c in (3, 5, 6, 7, 8, 9, 10):
                    cell.alignment = Alignment(horizontal="center")
                    
            row_num += 1

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(dest_path)

    @staticmethod
    def generate_assignment_excel(dest_path: str, header: dict, data_rows: list):
        """Generates the assignment Excel with exact column order as specified.

        Columns:
        ID | DESARROLLO | P.A. | CLIENTE | MZA | LOTE | EXT | INT |
        No.OFICIAL | FECHA INGRESO A RPP | AVISO | CLG |
        CANCELACION PRIMER AVISO | CANCELACION SEGUNDO AVISO | COMENTARIOS

        Each row groups all references of the same location (client+mz+lote+edif+viv).
        AVISO = first reference with concepto AVISO/AVISO PREVENTIVO
        CLG   = first reference with concepto CLG
        For RESERVADA: AVISO/CLG columns show reference numbers (referencia_portal).
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asignaciones"

        # ── Styles ──────────────────────────────────────────────────────────
        font_title    = Font(name="Calibri", size=14, bold=True)
        font_subtitle = Font(name="Calibri", size=11, bold=True, italic=True)
        font_header   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_data     = Font(name="Calibri", size=10)
        fill_header   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_alt      = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        border_thin   = Side(border_style="thin", color="BDD7EE")
        border_data   = Border(left=border_thin, right=border_thin,
                               top=border_thin, bottom=border_thin)

        # ── Column headers (row 1) ───────────────────────────────────────────
        COLUMNS = [
            "ID", "DESARROLLO", "P.A.", "CLIENTE",
            "MZA", "LOTE", "EXT", "INT",
            "No.OFICIAL", "FECHA REPORTA LA NOTARIA",
            "FECHA INGRESO A RPP", "FECHA ESCRITURA",
            "FECHA TITULACION", "AVISO", "CLG",
            "CANCELACION PRIMER AVISO", "CANCELACION SEGUNDO AVISO",
            "COMENTARIOS",
        ]
        ALIAS_AVISO = {"AVISO", "AVISO PREVENTIVO", "NUEVO_DERECHO_AVISO"}
        ALIAS_CLG   = {"CLG"}
        ALIAS_CANC1 = {"CANC_1ER _AVISO", "CANCELACION_1ER_AVISO"}
        ALIAS_CANC2 = {"CANC_2DO_AVISO", "CANCELACION_2DO_AVISO"}

        ws.row_dimensions[1].height = 30
        for ci, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_data

        # ── Pivot data by location / sequential index ────────────────────────
        # For ASIGNADA: group by client + location coordinates
        # For RESERVADA: pair the N-th AVISO with the N-th CLG so they share a row.
        pivot: dict = {}
        
        # Lists to keep track of RESERVADA references by concept type
        reservada_avisos = []
        reservada_clgs = []
        reservada_otros = []

        for r in data_rows:
            estado_r = (r.get("estado") or "").upper()
            is_reservada = (estado_r == "RESERVADA" or not any([
                r.get("cliente", ""), r.get("mz", ""), r.get("viv", "")
            ]))
            
            if is_reservada:
                concepto = (r.get("concepto") or "").strip().upper()
                if concepto in ALIAS_AVISO:
                    reservada_avisos.append(r)
                elif concepto in ALIAS_CLG:
                    reservada_clgs.append(r)
                else:
                    reservada_otros.append(r)
            else:
                key = (
                    r.get("cliente", ""),
                    r.get("desarrollo", ""),
                    r.get("mz", ""),
                    r.get("lote", ""),
                    r.get("edif", ""),
                    r.get("viv", ""),
                )
                if key not in pivot:
                    pivot[key] = {
                        "cliente":    r.get("cliente", ""),
                        "desarrollo": r.get("desarrollo", ""),
                        "mz":         r.get("mz", ""),
                        "lote_loc":   r.get("lote", ""),
                        "edif":       r.get("edif", ""),
                        "viv":        r.get("viv", ""),
                        "folio":      r.get("folio_electronico", ""),
                        "fecha_sol":  r.get("fecha_solicitud", ""),
                        "pa":         r.get("pa", ""),
                        "comentarios": r.get("comentarios", ""),
                        "fecha_reporte_notaria": r.get("fecha_reporte_notaria", ""),
                        "fecha_ingreso_rpp": r.get("fecha_ingreso_rpp", ""),
                        "fecha_escritura": r.get("fecha_escritura", ""),
                        "fecha_titulacion": r.get("fecha_titulacion", ""),
                        "aviso":      "",
                        "clg":        "",
                        "canc1":      "",
                        "canc2":      "",
                    }
                concepto = (r.get("concepto") or "").strip().upper()
                ref_val  = r.get("referencia", "")
                p = pivot[key]
                if concepto in ALIAS_AVISO and not p["aviso"]:
                    p["aviso"] = ref_val
                elif concepto in ALIAS_CLG and not p["clg"]:
                    p["clg"] = ref_val
                elif concepto in ALIAS_CANC1 and not p["canc1"]:
                    p["canc1"] = ref_val
                elif concepto in ALIAS_CANC2 and not p["canc2"]:
                    p["canc2"] = ref_val

        # Pair up the RESERVADA Avisos and CLGs sequentially
        max_pairs = max(len(reservada_avisos), len(reservada_clgs))
        for idx in range(max_pairs):
            aviso_ref = reservada_avisos[idx] if idx < len(reservada_avisos) else {}
            clg_ref = reservada_clgs[idx] if idx < len(reservada_clgs) else {}
            
            # Merge details using whichever has data
            ref_rep = aviso_ref if aviso_ref else clg_ref
            
            key = (f"__RESERVADA_PAIR_{idx}__", idx)
            pivot[key] = {
                "cliente":    ref_rep.get("cliente", "RESERVA PENDIENTE"),
                "desarrollo": ref_rep.get("desarrollo", ""),
                "mz":         ref_rep.get("mz", ""),
                "lote_loc":   ref_rep.get("lote", ""),
                "edif":       ref_rep.get("edif", ""),
                "viv":        ref_rep.get("viv", ""),
                "folio":      ref_rep.get("folio_electronico", ""),
                "fecha_sol":  ref_rep.get("fecha_solicitud", ""),
                "pa":         ref_rep.get("pa", ""),
                "comentarios": ref_rep.get("comentarios", ""),
                "fecha_reporte_notaria": ref_rep.get("fecha_reporte_notaria", ""),
                "fecha_ingreso_rpp": ref_rep.get("fecha_ingreso_rpp", ""),
                "fecha_escritura": ref_rep.get("fecha_escritura", ""),
                "fecha_titulacion": ref_rep.get("fecha_titulacion", ""),
                "aviso":      aviso_ref.get("referencia", ""),
                "clg":        clg_ref.get("referencia", ""),
                "canc1":      "",
                "canc2":      "",
            }
            
        # Append remaining non-Aviso/non-CLG RESERVADA references
        for idx, r in enumerate(reservada_otros):
            key = (f"__RESERVADA_OTRO_{idx}__", idx)
            concepto = (r.get("concepto") or "").strip().upper()
            ref_val  = r.get("referencia", "")
            
            pivot[key] = {
                "cliente":    r.get("cliente", "RESERVA PENDIENTE"),
                "desarrollo": r.get("desarrollo", ""),
                "mz":         r.get("mz", ""),
                "lote_loc":   r.get("lote", ""),
                "edif":       r.get("edif", ""),
                "viv":        r.get("viv", ""),
                "folio":      r.get("folio_electronico", ""),
                "fecha_sol":  r.get("fecha_solicitud", ""),
                "pa":         r.get("pa", ""),
                "comentarios": r.get("comentarios", ""),
                "fecha_reporte_notaria": r.get("fecha_reporte_notaria", ""),
                "fecha_ingreso_rpp": r.get("fecha_ingreso_rpp", ""),
                "fecha_escritura": r.get("fecha_escritura", ""),
                "fecha_titulacion": r.get("fecha_titulacion", ""),
                "aviso":      "",
                "clg":        "",
                "canc1":      ref_val if concepto in ALIAS_CANC1 else "",
                "canc2":      ref_val if concepto in ALIAS_CANC2 else "",
            }


        # ── Write data rows ──────────────────────────────────────────────────
        row_num = 2
        for idx, (key, p) in enumerate(pivot.items()):
            ws.row_dimensions[row_num].height = 18
            use_alt = (idx % 2 == 1)
            values = [
                idx + 1,            # ID = sequential row number
                p["desarrollo"],
                p["pa"],
                p["cliente"],
                p["mz"],
                p["lote_loc"],
                p["edif"] or "",
                p["viv"],
                p["folio"],                 # No.OFICIAL
                p["fecha_reporte_notaria"], # FECHA REPORTA LA NOTARIA
                p["fecha_ingreso_rpp"],     # FECHA INGRESO A RPP
                p["fecha_escritura"],       # FECHA ESCRITURA
                p["fecha_titulacion"],       # FECHA TITULACION
                p["aviso"],
                p["clg"],
                p["canc1"],
                p["canc2"],
                p.get("comentarios", ""),   # COMENTARIOS
            ]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font = font_data
                cell.border = border_data
                if use_alt:
                    cell.fill = fill_alt
                if ci in (5, 6, 7, 8, 10, 11, 12, 13):
                    cell.alignment = Alignment(horizontal="center")
            row_num += 1

        # ── Column widths ────────────────────────────────────────────────────
        col_widths = [10, 22, 12, 28, 8, 8, 8, 8, 18, 22, 22, 22, 22, 22, 22, 22, 22, 25]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # Freeze header rows
        ws.freeze_panes = "A2"

        wb.save(dest_path)


    @staticmethod
    def generate_blank_template(dest_path: str):
        """Generates a blank styled Excel sheet containing the correct columns for bulk import."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        font_header = Font(name="Calibri", size=10, bold=True, color="000000")
        fill_header = PatternFill(start_color="86C232", end_color="86C232", fill_type="solid") # Sleek green
        border_thin = Side(border_style="thin", color="D3D3D3")

        headers = [
            "RAZON SOCIAL",
            "DESARROLLO",
            "P.A",
            "CLIENTE",
            "MZA",
            "LOTE",
            "EXT",
            "INT",
            "No.OFICIAL",
            "FECHA REPORTA LA NOTARIA",
            "FECHA INGRESO A RPP",
            "FECHA ESCRITURA",
            "FECHA DE TITULACION",
            "ANALISIS",
            "AVISO",
            "CLG",
            "CANCELACION PRIMER AVISO",
            "CANCELACION SEGUNDO AVISO",
            "COMENTARIOS"
        ]
        
        ws.row_dimensions[1].height = 25
        for col_idx, text_header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=text_header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=Side(border_style="medium", color="000000"))

        # Add a mock row as a reference example for users
        ws.row_dimensions[2].height = 18
        mock_data = [
            "PROMOTORA RESIDENCIAL", "ALDEA TULUM", "1234", "JUAN PEREZ", "10", "5", "A", "3", "123456",
            "2026-07-20", "2026-07-21", "2026-07-22", "2026-07-25",
            "", "70028350819888101", "", "", "", "Sin observaciones"
        ]
        for col_idx, val in enumerate(mock_data, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10, italic=True)
            cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(dest_path)

