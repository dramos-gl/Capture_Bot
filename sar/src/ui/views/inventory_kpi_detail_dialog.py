"""Dialog to show complete detail of KPI stat cards from the Inventory module with full Excel export."""

import os
import datetime
from typing import Optional, List, Dict, Any

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QWidget, QFrame,
    QLabel, QLineEdit, QPushButton, QFileDialog
)
from PySide6.QtCore import Qt, QThread, Signal, QSize, QTimer
from PySide6.QtGui import QColor, QAction

from sar.src.ui.design_system.components.atoms.gl_label import CustomLabel
from sar.src.ui.design_system.components.atoms.gl_button import CustomButton
from sar.src.ui.design_system.components.molecules.gl_loading_dialog import GLLoadingDialog
from sar.src.ui.design_system.components.molecules.gl_labeled_combo import LabeledComboBox
from sar.src.ui.design_system.components.molecules.gl_menu import KeepOpenMenu
from sar.src.ui.design_system.components.organisms.gl_data_table import StyledDataTable
from sar.src.ui.design_system.components.organisms.gl_message_dialog import GLMessageBox as QMessageBox
from sar.src.ui.design_system.tokens.colors import Colors
from sar.src.ui.design_system.theme_manager import ThemeManager
from sar.src.ui.design_system.utils.icons import Icons
from sar.src.ui.design_system.utils.formatters import format_orden_filter_label


class KPIDetailLoadWorker(QThread):
    """Background worker thread to load KPI drill-down references from the DB."""
    result_ready = Signal(list, int, dict)
    error_occurred = Signal(str)

    def __init__(
        self, inventario_ui_service, filter_assigned: str,
        concepto_id: Optional[int], rfc_id: Optional[int],
        orden_ids: Optional[list], search_text: str = "",
        start_date: Optional[str] = None, end_date: Optional[str] = None
    ):
        super().__init__()
        self.inventario_ui_service = inventario_ui_service
        self.filter_assigned = filter_assigned
        self.concepto_id = concepto_id
        self.rfc_id = rfc_id
        self.orden_ids = orden_ids
        self.search_text = search_text
        self.start_date = start_date
        self.end_date = end_date
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def run(self):
        try:
            if self._is_cancelled:
                return
            # Load all matching records up to 10,000 for drill-down view
            res = self.inventario_ui_service.get_referencias_facturadas_paginated(
                limit=10000,
                offset=0,
                search_text=self.search_text,
                concepto_id=self.concepto_id,
                rfc_id=self.rfc_id,
                filter_assigned=self.filter_assigned,
                start_date=self.start_date,
                end_date=self.end_date,
                orden_ids=self.orden_ids
            )
            if self._is_cancelled:
                return
            summary = self.inventario_ui_service.get_inventario_summary(
                search_text=self.search_text,
                concepto_id=self.concepto_id,
                rfc_id=self.rfc_id,
                start_date=self.start_date,
                end_date=self.end_date,
                orden_ids=self.orden_ids
            )
            if not self._is_cancelled:
                self.result_ready.emit(res.get("records", []), res.get("total_count", 0), summary)
        except Exception as e:
            if not self._is_cancelled:
                self.error_occurred.emit(str(e))


class KPIDetailExcelWorker(QThread):
    """Background worker thread to build and save styled Excel workbook without freezing UI."""
    finished_success = Signal(str)
    error_occurred = Signal(str)

    def __init__(
        self, save_path: str, filtered_records: list,
        title_text: str, rfc_nombre: str, concepto_nombre: str,
        desarrollo_nombre: str = "Todos los desarrollos",
        delegacion_nombre: str = "Todas las delegaciones",
        destino_nombre: str = "Todos los destinos"
    ):
        super().__init__()
        self.save_path = save_path
        self.filtered_records = filtered_records
        self.title_text = title_text
        self.rfc_nombre = rfc_nombre
        self.concepto_nombre = concepto_nombre
        self.desarrollo_nombre = desarrollo_nombre
        self.delegacion_nombre = delegacion_nombre
        self.destino_nombre = destino_nombre

    def run(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detalle de Derechos"
            ws.views.sheetView[0].showGridLines = True

            # Colors & Styles
            blue_header = "2563EB"
            gray_zebra = "F8FAFC"
            border_gray = "CBD5E1"

            font_title = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
            font_sub = Font(name="Segoe UI", size=10, italic=True, color="64748B")
            font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=9, color="0F172A")
            font_data_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")

            fill_header = PatternFill(start_color=blue_header, end_color=blue_header, fill_type="solid")
            fill_zebra = PatternFill(start_color=gray_zebra, end_color=gray_zebra, fill_type="solid")
            fill_kpi = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

            thin_side = Side(border_style="thin", color=border_gray)
            border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            # Title Block
            ws["A1"] = f"SISTEMA DE ADMINISTRACIÓN DE REFERENCIAS (SAR) — {self.title_text.upper()}"
            ws["A1"].font = font_title
            
            filter_summary = (
                f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
                f"Empresa: {self.rfc_nombre} | Concepto: {self.concepto_nombre} | "
                f"Desarrollo: {self.desarrollo_nombre} | Delegación: {self.delegacion_nombre} | "
                f"Destino: {self.destino_nombre}"
            )
            ws["A2"] = filter_summary
            ws["A2"].font = font_sub

            # Summary KPIs row
            ws["A4"] = "TOTAL REGISTROS:"
            ws["A4"].font = font_data_bold
            ws["B4"] = len(self.filtered_records)
            ws["B4"].font = font_data_bold
            ws["B4"].alignment = align_left

            ws["D4"] = "IMPORTE TOTAL:"
            ws["D4"].font = font_data_bold
            total_importe = sum([float(r.get("importe") or 0) for r in self.filtered_records if r.get("importe")])
            ws["E4"] = total_importe
            ws["E4"].font = font_data_bold
            ws["E4"].number_format = "$#,##0.00"

            for col in range(1, 30):
                cell = ws.cell(row=4, column=col)
                cell.fill = fill_kpi

            # Table Headers
            headers = [
                "#", "REFERENCIA PORTAL", "FOLIO ORDEN", "EMPRESA / RFC", "CONCEPTO",
                "DELEGACIÓN", "IMPORTE", "ESTADO", "INTENTO", "CLIENTE",
                "CRÉDITO TITULAR", "DESARROLLO", "MZA", "LOTE", "EXT (EDIF)",
                "INT (VIV)", "FOLIO ELECTRÓNICO", "NO. OFICIAL", "P.A.", "FECHA SOLICITUD",
                "FECHA REPORTE NOTARÍA", "FECHA INGRESO RPP", "FECHA ESCRITURA",
                "FECHA TITULACIÓN", "DESTINO / ASIGNADO A", "TIPO ASIGNACIÓN",
                "SOLICITANTE EXTERNO", "FECHA ASIGNACIÓN", "COMENTARIOS"
            ]

            header_row = 6
            ws.row_dimensions[header_row].height = 26
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_cell

            # Data Rows
            current_row = 7
            for idx, r in enumerate(self.filtered_records, 1):
                ws.row_dimensions[current_row].height = 20
                is_zebra = (idx % 2 == 0)
                
                estado_code = r.get("estado_codigo") or ("ASIGNADA" if r.get("asignada") else "FACTURADA")
                imp_val = float(r.get("importe") or 0.0) if r.get("importe") else 0.0

                row_values = [
                    idx,
                    r.get("referencia_portal", ""),
                    r.get("folio_orden", ""),
                    r.get("empresa", ""),
                    r.get("concepto", ""),
                    r.get("delegacion", ""),
                    imp_val,
                    estado_code,
                    r.get("intento", "") or 1,
                    r.get("cliente", ""),
                    r.get("credito_titular", ""),
                    r.get("desarrollo", ""),
                    r.get("mz", ""),
                    r.get("lote", ""),
                    r.get("edif", ""),
                    r.get("viv", ""),
                    r.get("folio_electronico", ""),
                    r.get("no_oficial", ""),
                    r.get("pa", ""),
                    r.get("fecha_solicitud", ""),
                    r.get("fecha_reporte_notaria", ""),
                    r.get("fecha_ingreso_rpp", ""),
                    r.get("fecha_escritura", ""),
                    r.get("fecha_titulacion", ""),
                    r.get("asignado_a") or r.get("procesado_por") or "Sin Asignar",
                    r.get("tipo_asignacion", ""),
                    r.get("solicitante_externo", ""),
                    r.get("fecha_asignacion", ""),
                    r.get("comentarios", "")
                ]

                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = font_data
                    cell.border = border_cell
                    if is_zebra:
                        cell.fill = fill_zebra

                    if col_idx in (1, 3, 9, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 28):
                        cell.alignment = align_center
                    elif col_idx == 7:
                        cell.alignment = align_right
                        cell.number_format = "$#,##0.00"
                    elif col_idx in (2, 8):
                        cell.alignment = align_center
                        cell.font = font_data_bold
                    else:
                        cell.alignment = align_left

                current_row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < 6:
                        continue
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            wb.save(self.save_path)
            self.finished_success.emit(self.save_path)
        except Exception as e:
            self.error_occurred.emit(str(e))


class InventoryKPIDetailDialog(QDialog):
    """Rich modal dialog showing full drill-down table for KPI cards with multi-criteria filters and Excel export."""

    def __init__(
        self, db_connector, kpi_type: str,
        concepto_id: Optional[int] = None, concepto_nombre: str = "Todos los conceptos",
        rfc_id: Optional[int] = None, rfc_nombre: str = "Todas las empresas",
        orden_ids: Optional[list] = None, ordenes_count: int = 0,
        start_date: Optional[str] = None, end_date: Optional[str] = None,
        todas_las_ordenes: Optional[list] = None,
        parent=None
    ):
        super().__init__(parent)
        self.db_connector = db_connector
        from sar.src.services.inventario_ui_service import InventarioUIService
        self.inventario_ui_service = InventarioUIService(self.db_connector)
        
        self.kpi_type = kpi_type
        self.initial_concepto_id = concepto_id
        self.initial_concepto_nombre = concepto_nombre
        self.initial_rfc_id = rfc_id
        self.initial_rfc_nombre = rfc_nombre
        self.selected_orden_ids = list(orden_ids or [])
        self.todas_las_ordenes = list(todas_las_ordenes or [])
        self.start_date = start_date
        self.end_date = end_date

        self.all_records: List[Dict[str, Any]] = []
        self.filtered_records: List[Dict[str, Any]] = []
        self.active_worker: Optional[KPIDetailLoadWorker] = None

        # Debounce timer para búsqueda en Detalle KPI (700 ms)
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(700)
        self._search_timer.timeout.connect(self._on_search_trigger)

        # Resolve Title, State Filter & Colors
        if self.kpi_type == "disponibles":
            self.title_text = "Detalle de Derechos Disponibles"
            self.state_filter = "Disponible"
            self.header_color = Colors.PRIMARY
            self.icon_name = "clock"
        elif self.kpi_type == "asignadas":
            self.title_text = "Detalle de Derechos Asignados"
            self.state_filter = "Asignada"
            self.header_color = Colors.SUCCESS
            self.icon_name = "shield_check"
        elif self.kpi_type == "reservadas":
            self.title_text = "Detalle de Derechos Reservados"
            self.state_filter = "Reservada"
            self.header_color = "#F59E0B"
            self.icon_name = "archive"
        else: # "total"
            self.title_text = "Detalle de Total de Derechos"
            self.state_filter = "Todos"
            self.header_color = Colors.ACCENT
            self.icon_name = "file_text"

        self.setWindowTitle(self.title_text)
        self.resize(1380, 800)
        self.setMinimumSize(1150, 650)

        # Main Layout
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)

        # ── 1. Header Section ────────────────────────────────────────────────
        header_layout = QHBoxLayout()
        title_block = QVBoxLayout()
        title_block.setSpacing(2)
        
        self.lbl_title = CustomLabel(self.title_text, variant="header")
        self.lbl_title.setStyleSheet(f"color: {self.header_color}; font-size: 18px; font-weight: bold;")
        
        self.lbl_subtitle = CustomLabel("Filtros avanzados activos y vista detallada para auditoría y reportes", variant="body")
        self.lbl_subtitle.setStyleSheet("color: #64748B; font-size: 12px;")

        title_block.addWidget(self.lbl_title)
        title_block.addWidget(self.lbl_subtitle)
        header_layout.addLayout(title_block)
        header_layout.addStretch()

        # Close button in header
        btn_top_close = QPushButton("✕", self)
        btn_top_close.setFixedSize(30, 30)
        btn_top_close.setStyleSheet("""
            QPushButton {
                background: #F1F5F9;
                color: #64748B;
                border: 1px solid #E2E8F0;
                border-radius: 15px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #E2E8F0;
                color: #0F172A;
            }
        """)
        btn_top_close.clicked.connect(self.reject)
        header_layout.addWidget(btn_top_close)
        root.addLayout(header_layout)

        # ── 2. Metric Chips Bar ──────────────────────────────────────────────
        self.metric_frame = QFrame()
        self.metric_frame.setStyleSheet("""
            QFrame {
                background-color: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
            }
        """)
        metric_layout = QHBoxLayout(self.metric_frame)
        metric_layout.setContentsMargins(16, 8, 16, 8)
        metric_layout.setSpacing(24)

        self.lbl_metric_count = CustomLabel("Total Registros: 0", variant="body")
        self.lbl_metric_count.setStyleSheet("font-weight: bold; color: #1E293B;")

        self.lbl_metric_monto = CustomLabel("Importe Total: $0.00", variant="body")
        self.lbl_metric_monto.setStyleSheet("font-weight: bold; color: #1E293B;")

        self.lbl_metric_estado = CustomLabel(f"Filtro Estado: {self.state_filter.upper()}", variant="body")
        self.lbl_metric_estado.setStyleSheet(f"font-weight: bold; color: {self.header_color};")

        self.lbl_metric_ordenes = CustomLabel(f"Órdenes: {len(self.selected_orden_ids)} sel.", variant="body")
        self.lbl_metric_ordenes.setStyleSheet("color: #475569;")

        metric_layout.addWidget(self.lbl_metric_count)
        metric_layout.addWidget(self.lbl_metric_monto)
        metric_layout.addWidget(self.lbl_metric_estado)
        metric_layout.addWidget(self.lbl_metric_ordenes)
        metric_layout.addStretch()

        root.addWidget(self.metric_frame)

        # ── 3. Advanced Multi-Criteria Filter Bar (Horizontal, Uniform 36px) ───
        filter_bar_frame = QFrame(self)
        filter_bar_frame.setObjectName("filterBarFrame")
        filter_layout = QHBoxLayout(filter_bar_frame)
        filter_layout.setContentsMargins(12, 8, 12, 8)
        filter_layout.setSpacing(8)

        # 1. Search Box
        self.search_input = QLineEdit(self)
        self.search_input.setObjectName("filterBarSearch")
        self.search_input.setPlaceholderText("Buscar derecho, cliente, folio...")
        self.search_input.setMinimumWidth(220)
        self.search_input.setFixedHeight(36)
        self.search_input.setClearButtonEnabled(True)
        self.search_input.addAction(Icons.search("#64748B"), QLineEdit.LeadingPosition)
        self.search_input.textChanged.connect(self._on_search_text_changed)
        self.search_input.returnPressed.connect(self._on_search_trigger)
        filter_layout.addWidget(self.search_input, stretch=1)

        # Botón Buscar
        self.btn_buscar_kpi = QPushButton(self)
        self.btn_buscar_kpi.setObjectName("secondaryBtn")
        self.btn_buscar_kpi.setIcon(Icons.buscar("#FFFFFF") if ThemeManager.is_dark_active() else Icons.buscar("#334155"))
        self.btn_buscar_kpi.setFixedSize(36, 36)
        self.btn_buscar_kpi.setToolTip("Buscar (o presione Enter)")
        self.btn_buscar_kpi.clicked.connect(self._on_search_trigger)
        filter_layout.addWidget(self.btn_buscar_kpi)

        # 2. Combo Empresa
        self.labeled_empresa = LabeledComboBox("Empresa", ["Todas las empresas"])
        self.cb_empresa = self.labeled_empresa.combo
        self.cb_empresa.currentTextChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.labeled_empresa)

        # 3. Combo Concepto
        self.labeled_concepto = LabeledComboBox("Concepto", ["Todos los conceptos"])
        self.cb_concepto = self.labeled_concepto.combo
        self.cb_concepto.currentTextChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.labeled_concepto)

        # 4. Combo Desarrollo
        self.labeled_desarrollo = LabeledComboBox("Desarrollo", ["Todos los desarrollos"])
        self.cb_desarrollo = self.labeled_desarrollo.combo
        self.cb_desarrollo.currentTextChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.labeled_desarrollo)

        # 5. Combo Delegación
        self.labeled_delegacion = LabeledComboBox("Delegación", ["Todas las delegaciones"])
        self.cb_delegacion = self.labeled_delegacion.combo
        self.cb_delegacion.currentTextChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.labeled_delegacion)

        # 6. Combo Destino Asignado
        self.labeled_destino = LabeledComboBox("Destino Asignado", ["Todos los destinos", "NOTARIA", "COLABORADOR", "SIN ASIGNAR"])
        self.cb_destino = self.labeled_destino.combo
        self.cb_destino.currentTextChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.labeled_destino)

        # 7. Botón Filtro Órdenes (Embudo)
        self.btn_filter_orden = QPushButton(self)
        self.btn_filter_orden.setObjectName("secondaryBtn")
        self.btn_filter_orden.setIcon(Icons.filter_icon("#475569"))
        self.btn_filter_orden.setFixedSize(36, 36)
        self.btn_filter_orden.setToolTip("Filtrar por Órdenes")
        self.btn_filter_orden.clicked.connect(self._show_order_filter_menu)
        filter_layout.addWidget(self.btn_filter_orden)

        # 8. Refresh Button
        self.btn_refresh = QPushButton(self)
        self.btn_refresh.setObjectName("filterBarActionBtn")
        self.btn_refresh.setFixedSize(36, 36)
        self.btn_refresh.setIcon(Icons.actualizar("#FFFFFF"))
        self.btn_refresh.setIconSize(QSize(20, 20))
        self.btn_refresh.setToolTip("Recargar datos desde BD")
        self.btn_refresh.clicked.connect(self._load_data)
        filter_layout.addWidget(self.btn_refresh)

        # 9. Excel Button
        self.btn_excel = CustomButton("Exportar Excel", is_secondary=False)
        self.btn_excel.setIcon(Icons.file_excel("#FFFFFF"))
        self.btn_excel.setFixedHeight(36)
        self.btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #16A34A;
                color: #FFFFFF;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                padding: 0 14px;
            }
            QPushButton:hover {
                background-color: #15803D;
            }
        """)
        self.btn_excel.clicked.connect(self._on_export_excel)
        filter_layout.addWidget(self.btn_excel)

        root.addWidget(filter_bar_frame)

        # ── 4. Main Data Table with all asignacion_referencia fields ──────────
        headers = [
            "#", "Referencia", "Folio Orden", "Empresa", "Concepto",
            "Delegación", "Importe", "Estado", "Intento", "Cliente",
            "Crédito Titular", "Desarrollo", "Mz", "Lt", "Edif", "Viv",
            "Folio Electrónico", "No. Oficial", "P.A.", "Fecha Solicitud",
            "Fecha Reporte Notaría", "Fecha Ingreso RPP", "Fecha Escritura",
            "Fecha Titulación", "Destino / Asignado A", "Tipo", "Solicitante",
            "Fecha Asignación", "Comentarios"
        ]
        self.table = StyledDataTable(headers, parent=self)
        self.table.setMinimumHeight(350)
        root.addWidget(self.table)

        # ── 5. Footer Layout ─────────────────────────────────────────────────
        footer = QHBoxLayout()
        self.lbl_footer_info = CustomLabel("Cargando registros...", variant="muted")
        footer.addWidget(self.lbl_footer_info)
        footer.addStretch()

        btn_close = CustomButton("Cerrar", is_secondary=True)
        btn_close.setFixedHeight(36)
        btn_close.clicked.connect(self.accept)
        footer.addWidget(btn_close)

        root.addLayout(footer)

        # Trigger initial loading smoothly after dialog is shown
        QTimer.singleShot(50, self._load_data)

    def _populate_filter_dropdowns(self):
        """Populates the combo boxes dynamically from active records while maintaining selection."""
        cur_emp = self.cb_empresa.currentText()
        cur_con = self.cb_concepto.currentText()
        cur_des = self.cb_desarrollo.currentText()
        cur_del = self.cb_delegacion.currentText()

        empresas = sorted({str(r.get("empresa") or "").strip() for r in self.all_records if r.get("empresa")})
        conceptos = sorted({str(r.get("concepto") or "").strip() for r in self.all_records if r.get("concepto")})
        desarrollos = sorted({str(r.get("desarrollo") or "").strip() for r in self.all_records if r.get("desarrollo")})
        delegaciones = sorted({str(r.get("delegacion") or "").strip() for r in self.all_records if r.get("delegacion")})

        self.cb_empresa.blockSignals(True)
        self.cb_empresa.clear()
        self.cb_empresa.addItem("Todas las empresas")
        self.cb_empresa.addItems(empresas)
        if cur_emp in empresas:
            self.cb_empresa.setCurrentText(cur_emp)
        elif self.initial_rfc_nombre in empresas:
            self.cb_empresa.setCurrentText(self.initial_rfc_nombre)
        self.cb_empresa.blockSignals(False)

        self.cb_concepto.blockSignals(True)
        self.cb_concepto.clear()
        self.cb_concepto.addItem("Todos los conceptos")
        self.cb_concepto.addItems(conceptos)
        if cur_con in conceptos:
            self.cb_concepto.setCurrentText(cur_con)
        elif self.initial_concepto_nombre in conceptos:
            self.cb_concepto.setCurrentText(self.initial_concepto_nombre)
        self.cb_concepto.blockSignals(False)

        self.cb_desarrollo.blockSignals(True)
        self.cb_desarrollo.clear()
        self.cb_desarrollo.addItem("Todos los desarrollos")
        self.cb_desarrollo.addItems(desarrollos)
        if cur_des in desarrollos:
            self.cb_desarrollo.setCurrentText(cur_des)
        self.cb_desarrollo.blockSignals(False)

        self.cb_delegacion.blockSignals(True)
        self.cb_delegacion.clear()
        self.cb_delegacion.addItem("Todas las delegaciones")
        self.cb_delegacion.addItems(delegaciones)
        if cur_del in delegaciones:
            self.cb_delegacion.setCurrentText(cur_del)
        self.cb_delegacion.blockSignals(False)

    def _show_order_filter_menu(self):
        """Displays the popup menu for selecting orders."""
        if not self.todas_las_ordenes:
            QMessageBox.information(self, "Sin Órdenes", "No hay órdenes disponibles para filtrar.")
            return

        menu = KeepOpenMenu(self)
        order_actions = {}

        action_all = QAction("Todas las órdenes", menu, checkable=True)
        is_all_selected = len(self.selected_orden_ids) == len(self.todas_las_ordenes) and len(self.todas_las_ordenes) > 0
        action_all.setChecked(is_all_selected)

        def update_all_action_state():
            is_all = len(self.selected_orden_ids) == len(self.todas_las_ordenes) and len(self.todas_las_ordenes) > 0
            action_all.blockSignals(True)
            action_all.setChecked(is_all)
            action_all.blockSignals(False)

        def toggle_all(checked):
            if checked:
                self.selected_orden_ids = [ord["orden_id"] for ord in self.todas_las_ordenes]
            else:
                self.selected_orden_ids = []

            for oid, act in order_actions.items():
                act.blockSignals(True)
                act.setChecked(checked)
                act.blockSignals(False)

            self.lbl_metric_ordenes.setText(f"Órdenes: {len(self.selected_orden_ids)} sel.")
            self._load_data()

        action_all.triggered.connect(toggle_all)
        menu.addAction(action_all)
        menu.addSeparator()

        for ord in self.todas_las_ordenes:
            oid = ord["orden_id"]
            label = format_orden_filter_label(ord.get("folio", ""), ord.get("descripcion", ""))
            action = QAction(label, menu, checkable=True)
            action.setChecked(oid in self.selected_orden_ids)
            order_actions[oid] = action

            def make_toggle_handler(target_oid):
                def handler(checked):
                    if checked:
                        if target_oid not in self.selected_orden_ids:
                            self.selected_orden_ids.append(target_oid)
                    else:
                        if target_oid in self.selected_orden_ids:
                            self.selected_orden_ids.remove(target_oid)
                    update_all_action_state()
                    self.lbl_metric_ordenes.setText(f"Órdenes: {len(self.selected_orden_ids)} sel.")
                    self._load_data()
                return handler

            action.triggered.connect(make_toggle_handler(oid))
            menu.addAction(action)

        menu.exec(self.btn_filter_orden.mapToGlobal(self.btn_filter_orden.rect().bottomLeft()))

    def _load_data(self):
        """Asynchronously queries references from the database."""
        if self.active_worker and self.active_worker.isRunning():
            self.active_worker.cancel()
            self.active_worker.wait()

        self._loading_dialog = GLLoadingDialog("Cargando registros\ndel inventario...", self)

        self.active_worker = KPIDetailLoadWorker(
            inventario_ui_service=self.inventario_ui_service,
            filter_assigned=self.state_filter,
            concepto_id=None, # We load broad matching and filter in UI for maximum flexibility
            rfc_id=None,
            orden_ids=self.selected_orden_ids,
            search_text="",
            start_date=self.start_date,
            end_date=self.end_date
        )
        self.active_worker.result_ready.connect(self._on_data_loaded)
        self.active_worker.error_occurred.connect(self._on_data_error)
        self.active_worker.start()
        self._loading_dialog.exec()

    def _on_data_loaded(self, records: list, total_count: int, summary: dict):
        if hasattr(self, "_loading_dialog") and self._loading_dialog:
            self._loading_dialog.accept()

        self.all_records = records
        self._populate_filter_dropdowns()
        self._apply_filter_and_populate()

    def _on_data_error(self, error_msg: str):
        if hasattr(self, "_loading_dialog") and self._loading_dialog:
            self._loading_dialog.accept()
        QMessageBox.critical(self, "Error al Cargar Detalle", f"Ocurrió un error al consultar los datos:\n{error_msg}")

    def _on_search_text_changed(self, text: str):
        trimmed = text.strip()
        if not trimmed:
            if hasattr(self, "_search_timer"):
                self._search_timer.stop()
            self._on_search_trigger()
        else:
            if hasattr(self, "_search_timer"):
                self._search_timer.start()

    def _on_search_trigger(self):
        if hasattr(self, "_search_timer"):
            self._search_timer.stop()
        self._apply_filter_and_populate()

    def _on_filter_changed(self, _text: str = ""):
        self._apply_filter_and_populate()

    def _apply_filter_and_populate(self):
        query = self.search_input.text().strip().upper()
        emp_filter = self.cb_empresa.currentText()
        con_filter = self.cb_concepto.currentText()
        des_filter = self.cb_desarrollo.currentText()
        del_filter = self.cb_delegacion.currentText()
        dest_filter = self.cb_destino.currentText()

        self.filtered_records = []
        for r in self.all_records:
            # Check Empresa
            if emp_filter != "Todas las empresas" and (r.get("empresa") or "").strip() != emp_filter:
                continue
            # Check Concepto
            if con_filter != "Todos los conceptos" and (r.get("concepto") or "").strip() != con_filter:
                continue
            # Check Desarrollo
            if des_filter != "Todos los desarrollos" and (r.get("desarrollo") or "").strip() != des_filter:
                continue
            # Check Delegación
            if del_filter != "Todas las delegaciones" and (r.get("delegacion") or "").strip() != del_filter:
                continue
            # Check Destino Asignado
            if dest_filter == "NOTARIA":
                if str(r.get("tipo_asignacion") or "").upper() != "NOTARIA":
                    continue
            elif dest_filter == "COLABORADOR":
                if str(r.get("tipo_asignacion") or "").upper() != "COLABORADOR":
                    continue
            elif dest_filter == "SIN ASIGNAR":
                if r.get("asignada") or r.get("tipo_asignacion"):
                    continue

            # Text search filter
            if query:
                haystack = " ".join([
                    str(r.get("referencia_portal", "")),
                    str(r.get("folio_orden", "")),
                    str(r.get("empresa", "")),
                    str(r.get("concepto", "")),
                    str(r.get("delegacion", "")),
                    str(r.get("cliente", "")),
                    str(r.get("credito_titular", "")),
                    str(r.get("desarrollo", "")),
                    str(r.get("folio_electronico", "")),
                    str(r.get("no_oficial", "")),
                    str(r.get("pa", "")),
                    str(r.get("asignado_a", "")),
                    str(r.get("solicitante_externo", "")),
                    str(r.get("comentarios", "")),
                    str(r.get("intento", ""))
                ]).upper()
                if query not in haystack:
                    continue

            self.filtered_records.append(r)

        # Update metric chips
        total_monto = 0.0
        for r in self.filtered_records:
            try:
                imp = float(r.get("importe") or 0.0)
                total_monto += imp
            except (ValueError, TypeError):
                pass

        self.lbl_metric_count.setText(f"Total Registros: {len(self.filtered_records):,}")
        self.lbl_metric_monto.setText(f"Importe Total: ${total_monto:,.2f}")
        self.lbl_footer_info.setText(f"Mostrando {len(self.filtered_records)} de {len(self.all_records)} registros")

        # Populate table
        table_rows = []
        for idx, r in enumerate(self.filtered_records, 1):
            estado_code = r.get("estado_codigo") or ("ASIGNADA" if r.get("asignada") else "FACTURADA")
            imp_val = r.get("importe")
            imp_str = f"${float(imp_val):,.2f}" if imp_val else "$0.00"

            table_rows.append([
                str(idx),
                str(r.get("referencia_portal", "")),
                str(r.get("folio_orden", "")),
                str(r.get("empresa", "")),
                str(r.get("concepto", "")),
                str(r.get("delegacion", "")),
                imp_str,
                estado_code,
                str(r.get("intento", "") or "1"),
                str(r.get("cliente", "")),
                str(r.get("credito_titular", "")),
                str(r.get("desarrollo", "")),
                str(r.get("mz", "")),
                str(r.get("lote", "")),
                str(r.get("edif", "")),
                str(r.get("viv", "")),
                str(r.get("folio_electronico", "")),
                str(r.get("no_oficial", "")),
                str(r.get("pa", "")),
                str(r.get("fecha_solicitud", "")),
                str(r.get("fecha_reporte_notaria", "")),
                str(r.get("fecha_ingreso_rpp", "")),
                str(r.get("fecha_escritura", "")),
                str(r.get("fecha_titulacion", "")),
                str(r.get("asignado_a") or r.get("procesado_por") or "Sin Asignar"),
                str(r.get("tipo_asignacion", "")),
                str(r.get("solicitante_externo", "")),
                str(r.get("fecha_asignacion", "")),
                str(r.get("comentarios", ""))
            ])

        self.table.populate_rows(table_rows, checkable_first_col=False)

    def _on_export_excel(self):
        """Asynchronously generates an official styled Excel spreadsheet with the filtered records and displays a loading spinner."""
        if not self.filtered_records:
            QMessageBox.warning(self, "Sin Registros", "No hay registros disponibles para exportar.")
            return

        default_filename = f"Reporte_{self.kpi_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Reporte Excel", default_filename, "Archivos de Excel (*.xlsx)"
        )
        if not save_path:
            return

        self._export_loading_dialog = GLLoadingDialog("Generando y formateando\nreporte de Excel...", self)

        self._export_worker = KPIDetailExcelWorker(
            save_path=save_path,
            filtered_records=list(self.filtered_records),
            title_text=self.title_text,
            rfc_nombre=self.cb_empresa.currentText(),
            concepto_nombre=self.cb_concepto.currentText(),
            desarrollo_nombre=self.cb_desarrollo.currentText(),
            delegacion_nombre=self.cb_delegacion.currentText(),
            destino_nombre=self.cb_destino.currentText()
        )
        self._export_worker.finished_success.connect(self._on_export_success)
        self._export_worker.error_occurred.connect(self._on_export_error)
        self._export_worker.start()
        self._export_loading_dialog.exec()

    def _on_export_success(self, save_path: str):
        if hasattr(self, "_export_loading_dialog") and self._export_loading_dialog:
            self._export_loading_dialog.accept()
        QMessageBox.information(
            self, "Exportación Exitosa",
            f"El reporte se ha generado exitosamente en:\n{save_path}"
        )

    def _on_export_error(self, error_msg: str):
        if hasattr(self, "_export_loading_dialog") and self._export_loading_dialog:
            self._export_loading_dialog.accept()
        QMessageBox.critical(self, "Error al Exportar", f"No se pudo guardar el archivo Excel:\n{error_msg}")
