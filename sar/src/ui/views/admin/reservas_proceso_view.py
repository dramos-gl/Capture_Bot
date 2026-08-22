"""Bulk manual reservation view in administration Processes."""

import os
import openpyxl
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QMessageBox, QFileDialog, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QComboBox, QTextEdit, QLabel
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QBrush
from sar.src.ui.design_system.components.atoms.gl_button import CustomButton
from sar.src.ui.design_system.components.atoms.gl_label import CustomLabel
from sar.src.ui.design_system.utils.icons import Icons
from sqlalchemy import select
from sar.src.storage.models import Colaborador, Referencia

class ReservasProcesoView(QWidget):
    """View to manage manual bulk reservation of references for a selected Collaborator."""

    def __init__(self, db_connector, current_user_id, current_sesion_id, can_edit, parent=None):
        super().__init__(parent)
        self.db_connector = db_connector
        self.current_user_id = current_user_id
        self.current_sesion_id = current_sesion_id
        self.can_edit = can_edit
        self.payload_validado = [] # Loaded references ready to be saved

        # Main view layout
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(24, 24, 24, 24)
        self.layout.setSpacing(20)

        self._build_ui()
        self.refresh_data()

    def _build_ui(self):
        # 1. Main Header (Remove redundant "Administración del Sistema" title, show only sublabel)
        subtitle_layout = QHBoxLayout()
        subtitle_layout.setSpacing(6)
        sub_lbl = CustomLabel("Reserva Masiva de Referencias (Proceso Manual)", variant="subheader")
        subtitle_layout.addWidget(sub_lbl)
        
        info_icon = Icons.info() if hasattr(Icons, "info") else None
        if info_icon:
            info_lbl = QLabel()
            info_lbl.setPixmap(info_icon.pixmap(14, 14))
            subtitle_layout.addWidget(info_lbl)
        subtitle_layout.addStretch()
        self.layout.addLayout(subtitle_layout)

        # 2. Card 1: Configuration Form Container
        self.form_card = QFrame()
        self.form_card.setObjectName("cardFrame")
        self.form_card.setStyleSheet("""
            QFrame#cardFrame {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
            }
        """)
        
        form_card_layout = QVBoxLayout(self.form_card)
        form_card_layout.setContentsMargins(20, 20, 20, 20)
        form_card_layout.setSpacing(16)

        # Content Row (Colaborador on the left, Observations on the right)
        row_layout = QHBoxLayout()
        row_layout.setSpacing(24)

        # Left Column (Colaborador)
        left_col = QVBoxLayout()
        left_col.setSpacing(6)
        
        colab_title_lbl = QLabel()
        colab_title_lbl.setText("Colaborador Destino <font color='#EF4444'>*</font>")
        colab_title_lbl.setStyleSheet("font-weight: bold; font-size: 13px; color: #1E293B;")
        left_col.addWidget(colab_title_lbl)

        self.cb_colaborador = QComboBox()
        self.cb_colaborador.setFixedWidth(300)
        self.cb_colaborador.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border-radius: 6px;
                border: 1px solid #CBD5E1;
                background-color: #F8FAFC;
                font-size: 13px;
                color: #334155;
            }
        """)
        left_col.addWidget(self.cb_colaborador)

        colab_help_lbl = CustomLabel("Seleccione el colaborador que recibirá las referencias.", variant="muted")
        left_col.addWidget(colab_help_lbl)
        left_col.addStretch()
        
        row_layout.addLayout(left_col)

        # Right Column (Observations & Char Count)
        right_col = QVBoxLayout()
        right_col.setSpacing(6)
        
        obs_title_lbl = CustomLabel("Observaciones del Lote", variant="body")
        obs_title_lbl.setStyleSheet("font-weight: bold; font-size: 13px; color: #1E293B;")
        right_col.addWidget(obs_title_lbl)

        self.txt_observaciones = QTextEdit()
        self.txt_observaciones.setFixedHeight(80)
        self.txt_observaciones.setPlaceholderText("Ingrese el motivo o comentarios de esta reserva manual...")
        self.txt_observaciones.setStyleSheet("""
            QTextEdit {
                padding: 10px;
                border-radius: 6px;
                border: 1px solid #CBD5E1;
                background-color: #FFFFFF;
                font-size: 13px;
                color: #334155;
            }
        """)
        self.txt_observaciones.textChanged.connect(self._on_txt_changed)
        right_col.addWidget(self.txt_observaciones)

        char_lbl_layout = QHBoxLayout()
        char_lbl_layout.addStretch()
        self.lbl_char_counter = CustomLabel("0 / 500", variant="muted")
        char_lbl_layout.addWidget(self.lbl_char_counter)
        right_col.addLayout(char_lbl_layout)

        row_layout.addLayout(right_col)
        form_card_layout.addLayout(row_layout)

        # Form Actions Row (Separate select excel vs confirm reservation actions)
        actions_row = QHBoxLayout()
        actions_row.setSpacing(12)
        
        self.btn_descargar = CustomButton("Descargar Plantilla", is_secondary=True, icon_name="document")
        self.btn_descargar.setFixedWidth(160)
        actions_row.addWidget(self.btn_descargar)

        self.btn_seleccionar_excel = CustomButton("Seleccionar Excel", is_secondary=True, icon_name="search")
        self.btn_seleccionar_excel.setFixedWidth(160)
        actions_row.addWidget(self.btn_seleccionar_excel)

        self.btn_confirmar_reserva = CustomButton("Confirmar Reserva", icon_name="check")
        self.btn_confirmar_reserva.setFixedWidth(180)
        self.btn_confirmar_reserva.setEnabled(False) # Activated only when valid references are parsed
        actions_row.addWidget(self.btn_confirmar_reserva)

        self.btn_cancelar = CustomButton("Cancelar", is_secondary=True, icon_name="close")
        self.btn_cancelar.setFixedWidth(110)
        self.btn_cancelar.setStyleSheet("QPushButton#secondaryBtn { color: #EF4444; border-color: #FCA5A5; } QPushButton#secondaryBtn:hover { background-color: #FEF2F2; }")
        actions_row.addWidget(self.btn_cancelar)
        
        actions_row.addStretch()
        form_card_layout.addLayout(actions_row)

        self.layout.addWidget(self.form_card)

        # 3. Card 2: Results Container Card
        self.results_card = QFrame()
        self.results_card.setObjectName("cardFrame")
        self.results_card.setStyleSheet("""
            QFrame#cardFrame {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
            }
        """)
        
        results_card_layout = QVBoxLayout(self.results_card)
        results_card_layout.setContentsMargins(20, 20, 20, 20)
        results_card_layout.setSpacing(16)

        # Results header row (Title + Refresh button)
        results_header = QHBoxLayout()
        
        res_title_layout = QHBoxLayout()
        res_title_layout.setSpacing(8)
        table_icon = Icons.document() if hasattr(Icons, "document") else None
        if table_icon:
            table_icon_lbl = QLabel()
            table_icon_lbl.setPixmap(table_icon.pixmap(16, 16))
            res_title_layout.addWidget(table_icon_lbl)
        
        res_title_lbl = CustomLabel("Referencias Pre-cargadas / Estado de Validación", variant="subheader")
        res_title_lbl.setStyleSheet("font-weight: bold; font-size: 14px; color: #1E293B;")
        res_title_layout.addWidget(res_title_lbl)
        results_header.addLayout(res_title_layout)
        results_header.addStretch()

        self.btn_actualizar = CustomButton("Actualizar", is_secondary=True, icon_name="refresh")
        self.btn_actualizar.setFixedWidth(110)
        self.btn_actualizar.clicked.connect(self.refresh_data)
        results_header.addWidget(self.btn_actualizar)
        results_card_layout.addLayout(results_header)

        # Main Stacked Layout for Table or Placeholder
        self.preview_container = QFrame()
        self.preview_container.setStyleSheet("background-color: transparent; border: none;")
        self.preview_stack = QVBoxLayout(self.preview_container)
        self.preview_stack.setContentsMargins(0, 0, 0, 0)

        # Table Component
        self.table_res = QTableWidget()
        self.table_res.setColumnCount(3)
        self.table_res.setHorizontalHeaderLabels(["Referencia", "Estado de Validación", "Detalle de Validación"])
        self.table_res.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_res.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table_res.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_res.horizontalHeader().setStretchLastSection(True)
        self.table_res.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                gridline-color: #F1F5F9;
                background-color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #F8FAFC;
                padding: 10px;
                font-weight: bold;
                border: none;
                border-bottom: 1px solid #E2E8F0;
                color: #475569;
            }
        """)
        self.table_res.setVisible(False)
        self.preview_stack.addWidget(self.table_res)

        # Placeholder Component (No results to show)
        self.placeholder_widget = QWidget()
        self.placeholder_widget.setStyleSheet("background-color: transparent;")
        placeholder_layout = QVBoxLayout(self.placeholder_widget)
        placeholder_layout.setContentsMargins(40, 40, 40, 40)
        placeholder_layout.setSpacing(12)
        placeholder_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        search_icon = Icons.search() if hasattr(Icons, "search") else None
        if search_icon:
            search_icon_lbl = QLabel()
            search_icon_lbl.setPixmap(search_icon.pixmap(48, 48))
            search_icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            placeholder_layout.addWidget(search_icon_lbl)

        placeholder_title = CustomLabel("No hay referencias cargadas", variant="subheader")
        placeholder_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder_title.setStyleSheet("font-weight: bold; color: #475569;")
        placeholder_layout.addWidget(placeholder_title)

        placeholder_sub = CustomLabel("Seleccione un archivo Excel para validar las referencias.", variant="muted")
        placeholder_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder_layout.addWidget(placeholder_sub)

        self.preview_stack.addWidget(self.placeholder_widget)
        results_card_layout.addWidget(self.preview_container)

        # Table Footer (Pagination & Row Limits)
        self.footer_layout = QHBoxLayout()
        self.lbl_pagination = CustomLabel("Mostrando 0 a 0 de 0 resultados", variant="muted")
        self.footer_layout.addWidget(self.lbl_pagination)
        self.footer_layout.addStretch()

        self.footer_layout.addWidget(CustomLabel("Filas por página", variant="muted"))
        self.cb_page_limit = QComboBox()
        self.cb_page_limit.addItems(["25", "50", "100"])
        self.cb_page_limit.setFixedWidth(70)
        self.cb_page_limit.setStyleSheet("padding: 3px; border: 1px solid #CBD5E1; border-radius: 4px;")
        self.footer_layout.addWidget(self.cb_page_limit)

        self.btn_prev_page = CustomButton("<", is_secondary=True)
        self.btn_prev_page.setFixedWidth(30)
        self.btn_next_page = CustomButton(">", is_secondary=True)
        self.btn_next_page.setFixedWidth(30)
        self.footer_layout.addWidget(self.btn_prev_page)
        self.footer_layout.addWidget(self.btn_next_page)

        results_card_layout.addLayout(self.footer_layout)
        self.layout.addWidget(self.results_card)

        # Connect actions
        self.btn_descargar.clicked.connect(self._on_descargar_plantilla)
        self.btn_seleccionar_excel.clicked.connect(self._on_seleccionar_excel)
        self.btn_confirmar_reserva.clicked.connect(self._on_confirmar_reserva)
        self.btn_cancelar.clicked.connect(self._on_cancelar)

    def _on_cancelar(self):
        """Resets the UI form, loaded payload, observations, and restores placeholder state."""
        self.payload_validado = []
        self.btn_confirmar_reserva.setEnabled(False)
        self.txt_observaciones.clear()
        self.cb_colaborador.setCurrentIndex(0)
        self.table_res.setRowCount(0)
        self.table_res.setVisible(False)
        self.placeholder_widget.setVisible(True)
        self.lbl_pagination.setText("Mostrando 0 a 0 de 0 resultados")
        QMessageBox.information(self, "Proceso Cancelado", "Se han limpiado los datos cargados correctamente.")

    def _on_txt_changed(self):
        """Monitors and limits characters in the observations input."""
        txt = self.txt_observaciones.toPlainText()
        if len(txt) > 500:
            txt = txt[:500]
            self.txt_observaciones.setPlainText(txt)
            cursor = self.txt_observaciones.textCursor()
            cursor.movePosition(cursor.MoveOperation.End)
            self.txt_observaciones.setTextCursor(cursor)
        self.lbl_char_counter.setText(f"{len(txt)} / 500")

    def refresh_data(self):
        """Fetches active collaborators from database and populates the dropdown."""
        self.cb_colaborador.clear()
        self.cb_colaborador.addItem("-- Seleccione un Colaborador --", None)

        try:
            with self.db_connector.get_session() as session:
                stmt = select(Colaborador).where(Colaborador.activo == True).order_by(Colaborador.nombre)
                colabs = session.execute(stmt).scalars().all()
                for c in colabs:
                    self.cb_colaborador.addItem(c.nombre, c.colaborador_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar colaboradores: {e}")

    def _on_descargar_plantilla(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla de Reserva", 
            "plantilla_reserva_referencias.xlsx", 
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reserva Masiva"
            
            # Simple header columns
            ws.cell(row=1, column=1, value="REFERENCIA")
            ws.cell(row=1, column=2, value="ESTADO")

            # Sample row
            ws.cell(row=2, column=1, value="1234567890")
            ws.cell(row=2, column=2, value="RESERVADA")

            wb.save(file_path)
            QMessageBox.information(self, "Plantilla Guardada", "Se ha guardado la plantilla correctamente.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar la plantilla: {e}")

    def _on_seleccionar_excel(self):
        """Loads and pre-validates Excel contents, highlighting valid/invalid references in the table."""
        colab_id = self.cb_colaborador.currentData()
        if not colab_id:
            QMessageBox.warning(self, "Faltan Campos", "Por favor seleccione un colaborador de la lista antes de cargar el archivo.")
            return

        file_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel de Reserva", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                QMessageBox.warning(self, "Excel Vacío", "No se encontraron filas de datos para procesar.")
                return

            headers = [str(cell).strip().upper() for cell in rows[0] if cell is not None]
            if "REFERENCIA" not in headers or "ESTADO" not in headers:
                QMessageBox.critical(self, "Plantilla Inválida", "El archivo debe contener las columnas: REFERENCIA y ESTADO.")
                return

            ref_idx = headers.index("REFERENCIA")
            est_idx = headers.index("ESTADO")

            self.payload_validado = []
            results_to_render = []
            has_valid = False

            with self.db_connector.get_session() as session:
                for row in rows[1:]:
                    if len(row) <= max(ref_idx, est_idx):
                        continue
                    ref_val = str(row[ref_idx]).strip() if row[ref_idx] is not None else ""
                    est_val = str(row[est_idx]).strip().upper() if row[est_idx] is not None else "RESERVADA"

                    if not ref_val:
                        continue

                    # Validate reference existence in database
                    ref_db = session.execute(
                        select(Referencia).where(Referencia.referencia_portal == ref_val)
                    ).scalars().first()

                    if not ref_db:
                        results_to_render.append({
                            "referencia": ref_val,
                            "valido": False,
                            "msg": "ERROR: La referencia no existe en el sistema."
                        })
                    else:
                        has_valid = True
                        results_to_render.append({
                            "referencia": ref_val,
                            "valido": True,
                            "msg": f"VÁLIDA: Lista para asignación ({est_val})."
                        })
                        self.payload_validado.append({
                            "referencia_portal": ref_val,
                            "estado_codigo": est_val
                        })

            # Render validation results
            self.table_res.setRowCount(0)
            self.table_res.setRowCount(len(results_to_render))
            for idx, r in enumerate(results_to_render):
                item_ref = QTableWidgetItem(r["referencia"])
                item_status = QTableWidgetItem("🟢 VÁLIDO" if r["valido"] else "🔴 INVÁLIDO")
                item_msg = QTableWidgetItem(r["msg"])

                if not r["valido"]:
                    # Highlight invalid reference row with a light red background
                    item_ref.setBackground(QBrush(QColor("#FEE2E2")))
                    item_status.setBackground(QBrush(QColor("#FEE2E2")))
                    item_msg.setBackground(QBrush(QColor("#FEE2E2")))

                self.table_res.setItem(idx, 0, item_ref)
                self.table_res.setItem(idx, 1, item_status)
                self.table_res.setItem(idx, 2, item_msg)

            self.table_res.resizeColumnsToContents()
            self.placeholder_widget.setVisible(False)
            self.table_res.setVisible(True)
            
            self.lbl_pagination.setText(f"Mostrando 1 a {len(results_to_render)} de {len(results_to_render)} resultados")
            
            # Activate confirm button if there is at least one valid reference
            self.btn_confirmar_reserva.setEnabled(has_valid and self.can_edit)

            if has_valid:
                QMessageBox.information(
                    self, "Validación Completa", 
                    f"Se validó el archivo. Listas para guardar: {len(self.payload_validado)} referencias."
                )
            else:
                QMessageBox.warning(
                    self, "Sin referencias válidas", 
                    "No se encontró ninguna referencia válida en el archivo cargado."
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al procesar e importar la plantilla: {e}")

    def _on_confirmar_reserva(self):
        """Saves the pre-loaded valid references into the database after confirmation."""
        colab_id = self.cb_colaborador.currentData()
        colab_name = self.cb_colaborador.currentText()
        if not colab_id or not self.payload_validado:
            return

        # Confirmation Dialog
        reply = QMessageBox.question(
            self, "Confirmar Lote de Reserva",
            f"¿Está seguro de procesar y reservar {len(self.payload_validado)} referencias para el colaborador '{colab_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.No:
            return

        try:
            obs = self.txt_observaciones.toPlainText()

            # Save to Database
            from sar.src.storage.repositories import InventarioRepository
            with self.db_connector.get_session() as session:
                repo = InventarioRepository(session)
                result = repo.reservar_lote_manual_colaborador(
                    colaborador_id=colab_id,
                    observaciones=obs,
                    usuario_id=self.current_user_id,
                    referencias_estados=self.payload_validado
                )
                session.commit()

            # Success Dialog
            QMessageBox.information(
                self, "Proceso Exitoso",
                f"Lote de Reserva completado exitosamente.\nReferencias procesadas: {result['exitos']}"
            )

            # Clean UI and block confirm button
            self.payload_validado = []
            self.btn_confirmar_reserva.setEnabled(False)
            self.table_res.setRowCount(0)
            self.table_res.setVisible(False)
            self.placeholder_widget.setVisible(True)
            self.lbl_pagination.setText("Mostrando 0 a 0 de 0 resultados")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al confirmar la reserva en el sistema: {e}")
